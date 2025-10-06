import streamlit as st
import pandas as pd
import airportsdata
from geopy.distance import geodesic
import re
import folium
from streamlit_folium import st_folium
import os
import requests
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import tempfile
import plotly.express as px
from fuzzywuzzy import process
import plotly.graph_objects as go

# Initialize airports data
airports = airportsdata.load('IATA')

# ==================== CONFIGURATION ====================
GITHUB_GEO_MASTER_URL = "https://raw.githubusercontent.com/cxluo001/ghgcalculator/main/geo_master.xlsx"
GITHUB_EPA_EF_URL = "https://raw.githubusercontent.com/cxluo001/ghgcalculator/main/PAHO%20EF_usaepa_exiobase.xlsx"

# ==================== AWB Functions ====================
def clean_gross_weight(value):
    """Clean and standardize gross weight values"""
    if pd.isna(value):
        return None, None
    value_str = str(value).strip().upper()
    
    # Handle cases like '9320.000 K/Q', '352K 0', '659.000 K/Q', '1500 LB'
    # Extract numeric part and unit
    match = re.match(r'^([\d,.]+)\s*([A-Z]*)[^A-Z0-9]*$', value_str)
    if not match:
        # Try alternative patterns
        match = re.match(r'^([\d,.]+)\s*([A-Z]+)\s*[/\s]*[A-Z0-9]*$', value_str)
        if not match:
            return None, None
    
    numeric_part = match.group(1)
    unit_part = match.group(2) if len(match.groups()) > 1 else ''
    
    # Determine unit
    unit = None
    if any(u in unit_part for u in ['K', 'KG', 'KGS']):
        unit = 'KG'
    elif any(u in unit_part for u in ['LB', 'LBS', 'POUND']):
        unit = 'LB'  # Add pound unit recognition
    elif unit_part:
        unit = unit_part
        
    # Handle European decimal format
    is_european = False
    if ',' in numeric_part and '.' in numeric_part:
        if numeric_part.index('.') < numeric_part.index(','):
            is_european = True
    elif ',' in numeric_part:
        parts = numeric_part.split(',')
        if len(parts) == 2 and len(parts[1]) <= 3:
            is_european = True
            
    if is_european:
        numeric_part = numeric_part.replace('.', '').replace(',', '.')
    else:
        numeric_part = numeric_part.replace(',', '')
    
    try:
        weight = float(numeric_part)
        return weight, unit
    except ValueError:
        return None, None

def calculate_distances(flights):
    """Calculate distances for flight segments"""
    results = []
    total_distance = 0
    for i, flight in enumerate(flights, 1):
        row_data = {
            'leg': i,
            'origin': flight['origin'],
            'destination': flight['destination'],
            'airline': flight.get('airline', None),
            'origin_lat': None,
            'origin_lon': None,
            'dest_lat': None,
            'dest_lon': None,
            'distance_km': None
        }
        try:
            origin_coords = get_airport_coords(flight['origin'])
            dest_coords = get_airport_coords(flight['destination'])
            if None not in origin_coords and None not in dest_coords:
                distance = geodesic(origin_coords, dest_coords).kilometers
                row_data.update({
                    'origin_lat': origin_coords[0],
                    'origin_lon': origin_coords[1],
                    'dest_lat': dest_coords[0],
                    'dest_lon': dest_coords[1],
                    'distance_km': round(distance, 2)
                })
                total_distance += distance
        except Exception as e:
            print(f"Error calculating leg {i}: {e}")
        results.append(row_data)
    return results, round(total_distance, 2)

def parse_flight_route(route_str):
    """Extract flight segments from route string"""
    flights = []
    if pd.isna(route_str):
        return flights
    segments = [s.strip() for s in re.split(r'Flight \d+:', route_str) if s.strip()]
    for segment in segments:
        try:
            origin = re.search(r'Origin - ([^,]+)', segment).group(1).strip()
            dest = re.search(r'Destination - ([^,]+)', segment).group(1).strip()
            airline = re.search(r'Airline - ([^;]+)', segment)
            airline = airline.group(1).strip() if airline else None
            flights.append({'origin': origin, 'destination': dest, 'airline': airline})
        except (AttributeError, IndexError):
            print(f"Warning: Could not parse segment: {segment}")
    return flights

def get_airport_coords(location):
    """Get coordinates for airport code or city name"""
    if not location or pd.isna(location):
        return (None, None)
    location = str(location).upper().strip()
    if location in airports:
        return (airports[location]['lat'], airports[location]['lon'])
    matches = [ap for ap in airports.values() if ap['city'].upper() == location]
    if matches:
        return (matches[0]['lat'], matches[0]['lon'])
    clean_loc = re.sub(r'\b(AIRPORT|INTL|INTERNATIONAL|APT|ARPT)\b', '', location, flags=re.IGNORECASE).strip()
    if clean_loc != location:
        return get_airport_coords(clean_loc)
    return (None, None)

def get_spend_category_and_temp_control(apo, spend_df, reefer_df):
    """Get both spend category and temperature control status for a specific APO number"""
    try:
        # Convert APO to string and clean
        apo_str = str(apo).strip()
        print(f"Looking for APO: '{apo_str}'")  # Debug print
        
        # Check if spend_df is valid
        if spend_df is None or 'PurchaseOrderNumber' not in spend_df.columns:
            print("Spend dataframe is invalid or missing PurchaseOrderNumber column")
            return None, None
        
        # Convert spend_df PurchaseOrderNumber to string and clean
        spend_df = spend_df.copy()
        spend_df['PO_Clean'] = spend_df['PurchaseOrderNumber'].astype(str).str.strip()
        
        # Debug: Show sample of PO numbers in spend data
        print(f"Sample PO numbers in spend data: {spend_df['PO_Clean'].head(5).tolist()}")
        
        # Try exact match first
        exact_match = spend_df[spend_df['PO_Clean'] == apo_str]
        
        if not exact_match.empty:
            spend_category = exact_match['SpendCategory'].iloc[0]
            print(f"Exact match found for APO {apo_str}: '{spend_category}'")
        else:
            # Try various common formatting differences
            formats_to_try = [
                apo_str.replace('APO', '').replace('-', '').replace(' ', '').strip(),
                apo_str.replace('APO24-', '').replace('APO', '').strip(),
                apo_str.zfill(10),  # Pad with zeros
                apo_str.lstrip('0'),  # Remove leading zeros
                apo_str.upper(),
                apo_str.lower()
            ]
            
            # Try each format
            for fmt in formats_to_try:
                fmt_match = spend_df[spend_df['PO_Clean'] == fmt]
                if not fmt_match.empty:
                    spend_category = fmt_match['SpendCategory'].iloc[0]
                    print(f"Format match found for APO {apo_str} as '{fmt}': '{spend_category}'")
                    break
            else:
                # If no format worked, try partial matching
                partial_match = spend_df[spend_df['PO_Clean'].str.contains(apo_str, na=False)]
                if not partial_match.empty:
                    spend_category = partial_match['SpendCategory'].iloc[0]
                    print(f"Partial match found for APO {apo_str}: '{spend_category}'")
                else:
                    print(f"No match found for APO: {apo_str}")
                    return None, None
        
        if pd.isna(spend_category):
            print(f"Spend category is NaN for APO: {apo_str}")
            return None, None
        
        # Now get temperature control
        spend_category_clean = str(spend_category).strip().upper()
        print(f"Cleaned spend category: '{spend_category_clean}'")
        
        # Check if reefer_df is valid
        if reefer_df is None or 'SpendCategory' not in reefer_df.columns:
            print("Reefer dataframe is invalid")
            return spend_category, None
        
        # Standardize reefer dataframe
        reefer_df = reefer_df.copy()
        reefer_df['SpendCategory_Clean'] = reefer_df['SpendCategory'].astype(str).str.strip().str.upper()
        
        # Debug: Show sample categories in reefer data
        print(f"Sample reefer categories: {reefer_df['SpendCategory_Clean'].head(5).tolist()}")
        
        # Create mapping
        temp_control_map = reefer_df.drop_duplicates('SpendCategory_Clean').set_index('SpendCategory_Clean')['temp_control'].to_dict()
        
        # Get temperature control
        temperature_control = temp_control_map.get(spend_category_clean, None)
        print(f"Temperature control for '{spend_category_clean}': {temperature_control}")
        
        return spend_category, temperature_control
        
    except Exception as e:
        print(f"Error getting spend category/temp control for APO {apo}: {e}")
        import traceback
        print(traceback.format_exc())
        return None, None

def process_awb_file(file_path, sheet_name='AWB', spend_file=None, geo_content=None):
    """Process Excel file with AWB data and add temperature control + spend category"""
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=1)
        if 'Route/Stops' not in df.columns or 'Gross weight' not in df.columns:
            st.error("Required columns not found. Need both 'Route/Stops' and 'Gross weight'")
            return None
        
        # Load spend and geo data for temperature control if provided
        spend_df = None
        reefer_df = None
        
        if spend_file and geo_content:
            try:
                # Extract spend data
                spend_df = extract_spend_data(spend_file)
                # Extract temperature control mapping from geo_master
                geo_data = get_geo_master_data(geo_content)
                if geo_data[0] is not None:
                    _, _, _, reefer_df, _, _ = geo_data
            except Exception as e:
                st.warning(f"Could not load temperature control data: {e}")
        
        all_results = []
        for idx, row in df.iterrows():
            try:
                excel_row_num = idx + 3
                apo = row.iloc[0] if len(row) > 0 else None
                supplier = row.iloc[2] if len(row) > 2 else None
                
                if pd.isna(row['Route/Stops']):
                    continue
                    
                flights = parse_flight_route(row['Route/Stops'])
                if not flights:
                    continue
                
                leg_results, total_distance = calculate_distances(flights)
                gross_weight = row['Gross weight']
                clean_weight, uom = clean_gross_weight(gross_weight)
                
                gross_weight_ton = None
                if clean_weight is not None and uom is not None:
                    if uom.upper() in ['KG', 'KGS', 'K']:
                        gross_weight_ton = clean_weight / 1000
                    elif uom.upper() in ['LB', 'LBS', 'POUND']:
                        gross_weight_ton = clean_weight / 2204.62
                    else:
                        gross_weight_ton = clean_weight
                
                # Get spend category and temperature control if data is available
                spend_category = None
                temperature_control = None
                
                if spend_df is not None and reefer_df is not None and apo is not None:
                    spend_category, temperature_control = get_spend_category_and_temp_control(apo, spend_df, reefer_df)
                
                total_ghg_g = 0
                for result in leg_results:
                    result.update({
                        'awb_row': excel_row_num,
                        'APO': apo,
                        'Supplier': supplier,
                        'gross_weight': clean_weight,
                        'UoM': uom,
                        'gross_weight_ton': gross_weight_ton,
                        'Spend Category': spend_category,  # Add spend category
                        'Temperature Control': temperature_control  # Add temperature control
                    })
                    
                    if result['distance_km'] is not None:
                        result['E.F.'] = 1363 if result['distance_km'] < 1500 else 788
                        result['Units'] = 'g CO2e/t-km'
                        
                        if gross_weight_ton is not None:
                            ghg_g = result['E.F.'] * gross_weight_ton * result['distance_km']
                            result['ghg_emissions_gCO2e'] = ghg_g
                            result['ghg_emissions_tCO2e'] = ghg_g / 1000000
                            total_ghg_g += ghg_g
                    
                    all_results.append(result)
                
                all_results.append({
                    'awb_row': excel_row_num,
                    'APO': apo,
                    'Supplier': supplier,
                    'leg': 'TOTAL',
                    'origin': '',
                    'destination': '',
                    'airline': '',
                    'distance_km': None,
                    'gross_weight': None,
                    'UoM': None,
                    'gross_weight_ton': None,
                    'E.F.': None,
                    'Units': None,
                    'Spend Category': spend_category,  # Add to TOTAL row too
                    'Temperature Control': temperature_control,  # Add to TOTAL row too
                    'ghg_emissions_gCO2e': total_ghg_g if total_ghg_g != 0 else None,
                    'ghg_emissions_tCO2e': total_ghg_g / 1000000 if total_ghg_g != 0 else None
                })
                
            except Exception as e:
                print(f"Error processing row {idx + 3}: {e}")
                continue
        
        if not all_results:
            st.warning("No valid flight data found in file")
            return None
            
        results_df = pd.DataFrame(all_results)
        column_order = [
            'APO', 'Supplier', 'awb_row', 'leg', 'origin', 'destination', 'airline',
            'gross_weight', 'UoM', 'gross_weight_ton', 'distance_km',
            'E.F.', 'Units', 'ghg_emissions_gCO2e', 'ghg_emissions_tCO2e',
            'Spend Category', 'Temperature Control', 'origin_lat', 'origin_lon', 'dest_lat', 'dest_lon'
        ]
        column_order = [col for col in column_order if col in results_df.columns]
        
        return results_df[column_order]
        
    except Exception as e:
        st.error(f"Error processing file: {str(e)}")
        return None

def show_combined_map(awb_data, bol_data, precomputed_mappings):
    """Show combined air and ocean routes on map"""
    if (awb_data is None or awb_data.empty) and (bol_data is None or bol_data.empty):
        st.warning("No transport data available for mapping")
        return
    
    # Create base map
    avg_lat, avg_lon = 5, 0  # Default center if no data
    m = folium.Map(
        location=[avg_lat, avg_lon],
        zoom_start=2.4,
        tiles='CartoDB positron',
        control_scale=True
    )
    
    # Process AWB data (air routes - orange)
    if awb_data is not None and not awb_data.empty:
        awb_routes = awb_data.groupby([
            'origin', 'origin_lat', 'origin_lon',
            'destination', 'dest_lat', 'dest_lon'
        ]).agg({
            'APO': lambda x: ', '.join(sorted(set(x.astype(str)))) if len(x) > 0 else 'N/A',
            'Supplier': lambda x: ', '.join(sorted(set(x.astype(str)))) if len(x) > 0 else 'N/A',
            'distance_km': 'first',
            'ghg_emissions_tCO2e': 'sum',
            'leg': 'count'
        }).reset_index().rename(columns={'leg': 'count'})
        
        for _, route in awb_routes.iterrows():
            popup_content = f"""
            <div style="font-family: Arial; font-size: 14px; width: 250px">
                <h4 style="margin-bottom: 5px; color: #e67e22;">Air Route</h4>
                <hr style="margin: 5px 0;">
                <p style="margin: 3px 0;"><b>Route:</b> {route['origin']} → {route['destination']}</p>
                <p style="margin: 3px 0;"><b>Flights:</b> {route['count']}</p>
                <p style="margin: 3px 0;"><b>Distance:</b> {route['distance_km']:,.0f} km</p>
                <p style="margin: 3px 0;"><b>Total Emissions:</b> {route['ghg_emissions_tCO2e']:,.2f} tCO2e</p>
                <p style="margin: 3px 0;"><b>APOs:</b> {route['APO']}</p>
                <p style="margin: 3px 0;"><b>Suppliers:</b> {route['Supplier']}</p>
            </div>
            """
            
            folium.PolyLine(
                locations=[
                    [route['origin_lat'], route['origin_lon']],
                    [route['dest_lat'], route['dest_lon']]
                ],
                color='#e67e22',  # Orange for air
                weight=2,  # Reduced from 3 to make lines thinner
                opacity=0.2,  # Reduced from 0.7 to make more transparent
                popup=folium.Popup(popup_content, max_width=300),
                tooltip=f"Air: {route['origin']} → {route['destination']}"
            ).add_to(m)
    
    # Process BOL data (ocean routes - blue)
    if bol_data is not None and not bol_data.empty:
        # Use precomputed city_to_coords mapping
        city_to_coords = precomputed_mappings['city_to_coords']
        
        bol_routes = bol_data.groupby([
            'Port of loading', 'Port of discharge'
        ]).agg({
            'APO no.': lambda x: ', '.join(sorted(set(x.astype(str)))) if len(x) > 0 else 'N/A',
            'Shipper name': lambda x: ', '.join(sorted(set(x.astype(str)))) if len(x) > 0 else 'N/A',
            'Sea Distance (km)': 'first',
            'ghg_emissions_tCO2e': 'sum',
            'Measurement': 'count'
        }).reset_index().rename(columns={'Measurement': 'count'})
        
        for _, route in bol_routes.iterrows():
            origin = route['Port of loading'].upper()
            dest = route['Port of discharge'].upper()
            
            origin_coords = city_to_coords.get(origin)
            dest_coords = city_to_coords.get(dest)
            
            if origin_coords and dest_coords:
                popup_content = f"""
                <div style="font-family: Arial; font-size: 14px; width: 250px">
                    <h4 style="margin-bottom: 5px; color: #3498db;">Ocean Route</h4>
                    <hr style="margin: 5px 0;">
                    <p style="margin: 3px 0;"><b>Route:</b> {origin} → {dest}</p>
                    <p style="margin: 3px 0;"><b>Shipments:</b> {route['count']}</p>
                    <p style="margin: 3px 0;"><b>Distance:</b> {route['Sea Distance (km)']:,.0f} km</p>
                    <p style="margin: 3px 0;"><b>Total Emissions:</b> {route['ghg_emissions_tCO2e']:,.2f} tCO2e</p>
                    <p style="margin: 3px 0;"><b>APOs:</b> {route['APO no.']}</p>
                    <p style="margin: 3px 0;"><b>Suppliers:</b> {route['Shipper name']}</p>
                </div>
                """
                
                folium.PolyLine(
                    locations=[
                        [origin_coords['lat'], origin_coords['lng']],
                        [dest_coords['lat'], dest_coords['lng']]
                    ],
                    color='#3498db',  # Blue for ocean
                    weight=2,
                    opacity=0.6,
                    popup=folium.Popup(popup_content, max_width=300),
                    tooltip=f"Ocean: {origin} → {dest}"
                ).add_to(m)
    
    # Add legend
    legend_html = """
    <div style="position: fixed; bottom: 50px; left: 50px; width: 180px;
                z-index: 1000; background-color: white; padding: 10px;
                border: 2px solid grey; border-radius: 5px; font-size: 14px;
                font-family: Arial; box-shadow: 3px 3px 5px rgba(0,0,0,0.2)">
        <h4 style="margin: 0 0 8px 0; padding: 0;">Route Types</h4>
        <div style="display: flex; align-items: center; margin-bottom: 5px;">
            <div style="background: #e67e22; height: 20px; width: 20px; 
                        margin-right: 10px; opacity: 0.5;"></div>
            <span>Air Transport</span>
        </div>
        <div style="display: flex; align-items: center;">
            <div style="background: #3498db; height: 20px; width: 20px; 
                        margin-right: 10px; opacity: 0.4;"></div>
            <span>Ocean Transport</span>
        </div>
    </div>
    """
    m.get_root().html.add_child(folium.Element(legend_html))
    
    st_folium(m, width=1200, height=700, returned_objects=[])

def style_awb_dataframe(df):
    """Apply styling to highlight empty cells in flight legs"""
    def highlight_empty(val):
        if pd.isna(val) or val == '':
            return 'background-color: yellow'
        return ''
    
    # Apply styling only to flight legs (not TOTAL rows)
    styled_df = df.copy()
    if 'leg' in styled_df.columns:
        flight_legs_mask = styled_df['leg'].apply(lambda x: isinstance(x, int) or (isinstance(x, str) and x.isdigit()))
        styled_df = styled_df.style.applymap(highlight_empty, subset=pd.IndexSlice[flight_legs_mask, :])
    
    return styled_df

def create_excel_download(df, file_name):
    """Create an Excel file with highlighting for empty cells"""
    wb = Workbook()
    ws = wb.active
    
    # Create yellow fill for highlighting
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
    # Write dataframe to worksheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Highlight empty cells in flight legs (not TOTAL rows)
            if r_idx > 1:  # Skip header
                leg_value = df.iloc[r_idx-2]['leg'] if 'leg' in df.columns else None
                if (pd.isna(value) or value == '') and (isinstance(leg_value, int) or (isinstance(leg_value, str) and leg_value.isdigit())):
                    ws.cell(row=r_idx, column=c_idx).fill = yellow_fill
    
    # Save to temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        data = tmp.read()
    
    # Create download button
    st.download_button(
        label=f"Download {file_name}",
        data=data,
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    # Clean up
    os.unlink(tmp.name)

# ==================== BOL Functions ====================
def clean_bol_weight(value):
    """Clean and standardize BOL gross weight values"""
    if pd.isna(value):
        return None
    
    value_str = str(value).strip().upper()
    
    # Handle empty strings
    if not value_str:
        return None
    
    # Extract numeric part using regex that captures numbers with commas and decimals
    # This pattern handles: digits, commas, decimal points, and optional negative sign
    match = re.search(r'[-]?[\d,]+(?:\.\d+)?', value_str)
    if not match:
        return None
    
    numeric_part = match.group(0)
    
    # Handle European decimal format (comma as decimal separator)
    is_european = False
    if ',' in numeric_part and '.' in numeric_part:
        # If both comma and period are present, check which is the decimal separator
        comma_pos = numeric_part.find(',')
        dot_pos = numeric_part.find('.')
        if comma_pos > dot_pos:  # Format like "1.000,00" (European)
            is_european = True
    elif ',' in numeric_part and numeric_part.count(',') == 1:
        # Single comma - check if it's likely a decimal separator
        parts = numeric_part.split(',')
        if len(parts) == 2 and len(parts[1]) <= 3:
            is_european = True
            
    if is_european:
        numeric_part = numeric_part.replace('.', '').replace(',', '.')
    else:
        numeric_part = numeric_part.replace(',', '')
    
    try:
        return float(numeric_part)
    except ValueError:
        return None

def extract_bol_data(file_path, sheet_name='BoL'):
    """Extracts specific columns from the BoL worksheet"""
    target_columns = [
        'APO no.', 'Port of loading', 'Port of discharge',
        'Shipper name', 'Shipper address', 'Consignee',
        'Consignee address', 'Gross weight, kg', 'Measurement'
    ]
    
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=1)
        extracted_df = df[target_columns].copy()
        
        # Clean the Gross weight column using our new function
        extracted_df['Gross weight, kg'] = extracted_df['Gross weight, kg'].apply(clean_bol_weight)
        
        return extracted_df
    except Exception as e:
        st.error(f"Error processing BoL data: {e}")
        return None

@st.cache_data
def load_and_process_geo_master(_geo_content):
    """Load and preprocess geo master data with caching from GitHub content"""
    try:
        # _geo_content is bytes from GitHub - convert to BytesIO
        geo_file_obj = io.BytesIO(_geo_content)
        
        # Now geo_file_obj is always a BytesIO object with seek capability
        geo_df = pd.read_excel(
            geo_file_obj,
            sheet_name='geo',
            usecols=['city_ascii', 'lat', 'lng', 'country', 'iso3', 'continent']
        )
        geo_file_obj.seek(0)  # Reset to beginning
        
        cerdi_df = pd.read_excel(
            geo_file_obj,
            sheet_name='cerdi',
            usecols=['iso1', 'iso2', 'seadistance_km']
        )
        geo_file_obj.seek(0)
        
        ef_df = pd.read_excel(
            geo_file_obj,
            sheet_name='ef',
            usecols=['Port of Origin', 'Port of Destination', 
                    'Corresponding trade line group', 'Reefer', 'Dry']
        )
        geo_file_obj.seek(0)
        
        reefer_df = pd.read_excel(
            geo_file_obj,
            sheet_name='temperature',
            usecols=['SpendCategory', 'temp_control']
        )
        geo_file_obj.seek(0)
        
        supplier_df = pd.read_excel(
            geo_file_obj,
            sheet_name='supplier',
            usecols=['Supplier', 'country', 'supplier_specific_ef_sc123']
        )
        
        # Preprocess data for faster access
        # 1. Geo data: create optimized mappings
        geo_unique = geo_df.drop_duplicates(subset=['city_ascii'], keep='first').copy()
        geo_unique['city_ascii_upper'] = geo_unique['city_ascii'].str.upper()
        city_to_coords = {}
        for _, row in geo_unique.iterrows():
            city = row['city_ascii']
            if city not in city_to_coords:
                city_to_coords[city] = {'lat': row['lat'], 'lng': row['lng']}
        city_to_continent = geo_unique.set_index('city_ascii_upper')['continent'].to_dict()
        city_to_iso3 = geo_unique.set_index('city_ascii_upper')['iso3'].to_dict()
        
        # 2. CERDI data: create distance mapping
        distance_map = cerdi_df.set_index(['iso1', 'iso2'])['seadistance_km'].to_dict()
        
        # 3. EF data: create trade line mapping
        trade_map = ef_df.set_index(['Port of Origin', 'Port of Destination'])[
            ['Corresponding trade line group', 'Reefer', 'Dry']
        ].to_dict('index')
        
        # 4. Reefer data: create temperature control mapping
        reefer_df = reefer_df.copy()
        reefer_df['SpendCategory_Upper'] = reefer_df['SpendCategory'].str.strip().str.upper()
        temp_control_map = reefer_df.set_index('SpendCategory_Upper')['temp_control'].to_dict()
        
        # 5. Supplier data: create supplier mappings
        supplier_country_map = {}
        supplier_ef_map = {}
        for _, row in supplier_df.iterrows():
            if pd.notna(row['Supplier']):
                supplier_country_map[row['Supplier']] = row['country'] if pd.notna(row['country']) else None
                supplier_ef_map[row['Supplier']] = row['supplier_specific_ef_sc123'] if pd.notna(row['supplier_specific_ef_sc123']) else None
        
        return {
            'geo_df': geo_df,
            'cerdi_df': cerdi_df,
            'ef_df': ef_df,
            'reefer_df': reefer_df,
            'supplier_df': supplier_df,
            'precomputed_mappings': {
                'city_to_coords': city_to_coords,
                'city_to_continent': city_to_continent,
                'city_to_iso3': city_to_iso3,
                'distance_map': distance_map,
                'trade_map': trade_map,
                'temp_control_map': temp_control_map,
                'supplier_country_map': supplier_country_map,
                'supplier_ef_map': supplier_ef_map
            }
        }
        
    except Exception as e:
        st.error(f"Error processing geo_master.xlsx: {e}")
        import traceback
        st.error(f"Traceback: {traceback.format_exc()}")
        return None

def get_geo_master_data(geo_content):
    """Get processed geo master data with cached mappings from GitHub content"""
    processed_data = load_and_process_geo_master(geo_content)
    if processed_data is None:
        return None, None, None, None, None, None
    
    return (
        processed_data['geo_df'],
        processed_data['cerdi_df'], 
        processed_data['ef_df'],
        processed_data['reefer_df'],
        processed_data['supplier_df'],
        processed_data['precomputed_mappings']
    )

def extract_spend_data(file_path):
    """Extracts spend data with expanded columns"""
    target_columns = [
        "YearReceipt", "PurchaseOrderNumber", "Supplier", 
        "ShippingMethod", "SpendCategory", "LineDescription",
        "FundType", "ShipToAddressCountry", "Freight per APO", 
        "Amount per PO Line"
    ]
    try:
        spend_df = pd.read_excel(
            file_path,
            usecols=target_columns
        )
        spend_df["SpendCategory"] = spend_df["SpendCategory"].str.strip().str.upper()
        spend_df["FundType"] = spend_df["FundType"].str.strip()
        return spend_df
    except Exception as e:
        st.error(f"Error processing spend data: {e}")
        return None

def enrich_bol_with_geo(bol_df, geo_df):
    """Enriches bol_df with geo information"""
    try:
        geo_unique = geo_df.drop_duplicates(subset=['city_ascii'], keep='first')
        geo_unique['city_ascii'] = geo_unique['city_ascii'].str.upper()
        city_to_iso3 = geo_unique.set_index('city_ascii')['iso3'].to_dict()
        city_to_continent = geo_unique.set_index('city_ascii')['continent'].to_dict()
        
        bol_df['Port of loading'] = bol_df['Port of loading'].str.upper()
        bol_df['Port of discharge'] = bol_df['Port of discharge'].str.upper()

        bol_df['Loading Port ISO3'] = bol_df['Port of loading'].map(city_to_iso3)
        bol_df['Loading Port Continent'] = bol_df['Port of loading'].map(city_to_continent)
        bol_df['Discharge Port ISO3'] = bol_df['Port of discharge'].map(city_to_iso3)
        bol_df['Discharge Port Continent'] = bol_df['Port of discharge'].map(city_to_continent)
        
        return bol_df
    except Exception as e:
        st.error(f"Error enriching bol data: {e}")
        return bol_df

def add_sea_distance(bol_df, cerdi_df):
    """Adds sea distance information"""
    try:
        distance_map = cerdi_df.set_index(['iso1', 'iso2'])['seadistance_km'].to_dict()
        bol_df['Sea Distance (km)'] = bol_df.apply(
            lambda row: distance_map.get(
                (row['Loading Port ISO3'], row['Discharge Port ISO3']), 
                None
            ),
            axis=1
        )
        return bol_df
    except Exception as e:
        st.error(f"Error adding sea distance: {e}")
        return bol_df

def add_trade_line_info(bol_df, ef_df):
    """Adds trade line information"""
    try:
        trade_map = ef_df.set_index(['Port of Origin', 'Port of Destination'])[
            ['Corresponding trade line group', 'Reefer', 'Dry']
        ].to_dict('index')
        
        bol_df[['Trade Line Group', 'Reefer', 'Dry']] = bol_df.apply(
            lambda row: pd.Series(
                trade_map.get(
                    (row['Loading Port Continent'], row['Discharge Port Continent']),
                    {'Corresponding trade line group': None, 'Reefer': None, 'Dry': None}
                )
            ),
            axis=1
        )
        return bol_df
    except Exception as e:
        st.error(f"Error adding trade line information: {e}")
        return bol_df

def add_spend_category(bol_df, spend_df):
    """Adds SpendCategory information"""
    try:
        spend_category_map = spend_df.drop_duplicates('PurchaseOrderNumber').set_index('PurchaseOrderNumber')['SpendCategory'].to_dict()
        bol_df['SpendCategory'] = bol_df['APO no.'].map(spend_category_map)
        return bol_df
    except Exception as e:
        st.error(f"Error adding spend category: {e}")
        return bol_df

def add_temp_control(bol_df, reefer_df):
    """Adds temperature control information with case-insensitive matching"""
    try:
        # Create case-insensitive mapping
        reefer_df = reefer_df.copy()
        reefer_df['SpendCategory_Upper'] = reefer_df['SpendCategory'].str.strip().str.upper()
        
        temp_control_map = reefer_df.drop_duplicates('SpendCategory_Upper').set_index('SpendCategory_Upper')['temp_control'].to_dict()
        
        # Convert bol_df categories to uppercase for matching
        if 'SpendCategory' in bol_df.columns:
            bol_df = bol_df.copy()
            bol_df['SpendCategory_Upper'] = bol_df['SpendCategory'].str.strip().str.upper()
            bol_df['Temperature Control'] = bol_df['SpendCategory_Upper'].map(temp_control_map)
            bol_df = bol_df.drop('SpendCategory_Upper', axis=1)
        
        return bol_df
    except Exception as e:
        st.error(f"Error adding temperature control: {e}")
        return bol_df

def add_teu_column(bol_df):
    """Adds TEU column"""
    try:
        bol_df = bol_df.copy()
        bol_df['teu'] = bol_df['Gross weight, kg'] / 10000
        return bol_df
    except Exception as e:
        st.error(f"Error calculating TEU: {e}")
        return bol_df

def add_ghg_emissions(bol_df):
    """Adds GHG emissions columns"""
    try:
        bol_df = bol_df.copy()
        bol_df['ghg_emissions_gCO2e'] = bol_df.apply(
            lambda row: (row['teu'] * row['Sea Distance (km)'] * 
                        (row['Reefer'] if row['Temperature Control'] == 'YES' else row['Dry'])),
            axis=1
        )
        bol_df['ghg_emissions_tCO2e'] = bol_df['ghg_emissions_gCO2e'] / 1000000
        return bol_df
    except Exception as e:
        st.error(f"Error calculating GHG emissions: {e}")
        return bol_df

def process_bol_file(bol_file, geo_content, spend_file):
    """Process BOL data with comprehensive error handling"""
    try:
        # Load geo master data with new function
        geo_data = get_geo_master_data(geo_content)
        if geo_data[0] is None:
            st.error("Failed to load geo master data")
            return None
            
        geo_df, cerdi_df, ef_df, reefer_df, supplier_df, precomputed_mappings = geo_data
        
        # Load BOL data
        bol_df = extract_bol_data(bol_file)
        if bol_df is None or bol_df.empty:
            st.error("No valid BOL data found or failed to load BOL file")
            return None
            
        required_columns = ['Gross weight, kg', 'Port of loading', 'Port of discharge', 'APO no.']
        missing_cols = [col for col in required_columns if col not in bol_df.columns]
        if missing_cols:
            st.error(f"BOL file is missing required columns: {', '.join(missing_cols)}")
            return None
            
        # Load spend data
        spend_df = extract_spend_data(spend_file)
        if spend_df is None or spend_df.empty:
            st.error("Failed to load spend data")
            return None

        # Initialize processing
        final_bol_df = bol_df.copy()
        
        # Step 2: Data enrichment pipeline
        def safe_apply_step(df, step):
            try:
                missing = [col for col in step.get("required_cols", []) if col not in df.columns]
                if missing:
                    st.warning(f"Skipped {step['name']} - missing columns: {missing}")
                    return df
                result = step["func"](df.copy(), *step.get("deps", []))
                return result if result is not None else df
            except Exception as e:
                st.warning(f"Warning during {step['name']}: {str(e)}")
                return df

        enrichment_steps = [
            {"name": "Geo information", "func": enrich_bol_with_geo, "deps": [geo_df]},
            {"name": "Sea distance", "func": add_sea_distance, "deps": [cerdi_df], 
             "required_cols": ["Loading Port ISO3", "Discharge Port ISO3"]},
            {"name": "Trade line", "func": add_trade_line_info, "deps": [ef_df],
             "required_cols": ["Loading Port Continent", "Discharge Port Continent"]},
            {"name": "Spend category", "func": add_spend_category, "deps": [spend_df],
             "required_cols": ["APO no."]},
            {"name": "Temperature", "func": add_temp_control, "deps": [reefer_df],
             "required_cols": ["SpendCategory"]}
        ]

        for step in enrichment_steps:
            final_bol_df = safe_apply_step(final_bol_df, step)

        # Step 3: Calculations
        calculation_steps = [
            {
                "name": "TEU calculation",
                "func": lambda df: add_teu_column(df) if 'Gross weight, kg' in df.columns else df,
                "required_cols": ["Gross weight, kg"]
            },
            {
                "name": "GHG emissions",
                "func": lambda df: add_ghg_emissions(df) if all(col in df.columns for col in [
                    'teu', 'Sea Distance (km)', 'Temperature Control', 'Reefer', 'Dry'
                ]) else df,
                "required_cols": ["teu", "Sea Distance (km)", "Temperature Control", "Reefer", "Dry"]
            }
        ]

        for calc in calculation_steps:
            final_bol_df = safe_apply_step(final_bol_df, calc)

        # Final validation
        if final_bol_df.empty:
            st.error("Processing resulted in empty dataframe")
            return None
            
        if 'ghg_emissions_gCO2e' not in final_bol_df.columns:
            st.warning("GHG emissions calculation failed - missing required columns")
            
        return final_bol_df
        
    except Exception as e:
        st.error(f"Critical error processing BOL data: {str(e)}")
        return None

def calculate_scope31_emissions(spend_file, awb_data, bol_data, selected_years):
    """Calculate Scope 3.1 emissions with hardcoded EPA table and exact matching - ONLY for APOs in AWB/BOL"""
    try:
        # Hardcoded EPA data
        epa_data = [
            ["(PRO) Antiretroviral medicines", "Pharmaceutical Preparation Manufacturing", "325412", 0.045],
            ["(PRO) Antileishmaniasis", "Pharmaceutical Preparation Manufacturing", "325412", 0.045],
            ["(PRO) Antitubercular drugs", "Pharmaceutical Preparation Manufacturing", "325412", 0.045],
            ["(PRO) Disease vectors management and control", "Exterminating and Pest Control Services", "561700", 0.214],
            ["(PRO) Vaccines and antigens and toxoids", "Biological product (except diagnostics) manufacturing", "325414", 0.126],
            ["(PRO) Antibacterials", "Pharmaceutical Preparation Manufacturing", "325412", 0.045],
            ["(PRO) Anti-malarial", "Pharmaceutical Preparation Manufacturing", "325412", 0.045],
            ["(PRO) Antineoplastic agents", "Pharmaceutical Preparation Manufacturing", "325412", 0.045],
            ["(PRO) Cardiovascular Medicines", "Pharmaceutical Preparation Manufacturing", "325412", 0.045],
            ["(PRO) Nutritional Supplements", "All Other Miscellaneous Food Manufacturing", "311990", 0.358],
            ["(PRO) Antiviral drugs", "Pharmaceutical Preparation Manufacturing", "325412", 0.045],
            ["(PRO) Clinical and diagnostic analyzers and accessories and supplies", "In-Vitro Diagnostic Substance Manufacturing", "325413", 0.161],
            ["(PRO) Anti-Chagas", "Pharmaceutical Preparation Manufacturing", "325412", 0.045],
            ["(PRO) Antifungal drugs", "Pharmaceutical Preparation Manufacturing", "325412", 0.045],
            ["(PRO) Syringes and accessories", "Surgical and Medical Instrument Manufacturing", "339112", 0.119],
            ["(PRO) Rapid test kits", "In-Vitro Diagnostic Substance Manufacturing", "325413", 0.161],
            ["(PRO) Antipsychotics", "Pharmaceutical Preparation Manufacturing", "325412", 0.045],
            ["(PRO) Central nervous system drugs", "Pharmaceutical Preparation Manufacturing", "325412", 0.045],
            ["(PRO) Laboratory and scientific equipment", "Surgical and Medical Instrument Manufacturing", "339112", 0.119],
            ["(PRO) Needle or blade or sharps disposal container or cart", "Surgical and Medical Instrument Manufacturing", "339112", 0.119],
            ["(PRO) Veterinary vaccines and virology products", "Biological product (except diagnostics) manufacturing", "325414", 0.126],
            ["(PRO) Laboratory supplies and fixtures", "Surgical and Medical Instrument Manufacturing", "339112", 0.119],
            ["(PRO) Medical Equipment and Accessories and Supplies", "Surgical and Medical Instrument Manufacturing", "339112", 0.119],
            ["(CC) (PAHO internal use) Equipment Services: installation, maintenance, leasing and training", "In-Vitro Diagnostic Substance Manufacturing", "325413", 0.161],
            ["(PRO) Hematolic drugs", "Pharmaceutical Preparation Manufacturing", "325412", 0.045],
            ["(PRO) Anaesthetic drugs and related adjuncts and analeptics",  "Pharmaceutical Preparation Manufacturing", "325412", 0.045],
            ["(PRO) Patient care and treatment products and supplies",  "Surgical and Medical Instrument Manufacturing", "339112", 0.119],
            ["(PRO) Antidotes and emetics", "Pharmaceutical Preparation Manufacturing", "325412", 0.045],
            ["(PRO) Industrial freezers and refrigerators", "Air-Conditioning and Warm Air Heating Equipment and Commercial and Industrial Refrigeration Equipment Manufacturing", "333415", 0.156],
            ["(PRO) Antidiabetic agents and hyperglycemic agents", "Pharmaceutical Preparation Manufacturing", "325412", 0.045],
            ["(PRO) Cold storage box", "Air-Conditioning and Warm Air Heating Equipment and Commercial and Industrial Refrigeration Equipment Manufacturing", "333415", 0.156],
            ["(PRO) Cold pack or ice brick",  "Air-Conditioning and Warm Air Heating Equipment and Commercial and Industrial Refrigeration Equipment Manufacturing", "333415", 0.156],
            ["(CC) (Technical Cooperation) Supplies and Materials",  "Surgical and Medical Instrument Manufacturing", "339112", 0.119],
            ["(PRO) Drugs affecting the respiratory tract",  "Pharmaceutical Preparation Manufacturing", "325412", 0.045],
            ["(PRO) Electrolytes",  "Pharmaceutical Preparation Manufacturing", "325412", 0.045],
            ["(PRO)Temperature and heat measuring instruments",  "Surgical and Medical Instrument Manufacturing", "339112", 0.119],
            ["(PRO) Muscle Relaxant Medicines",  "Pharmaceutical Preparation Manufacturing", "325412", 0.045],
            ["(CC) (Technical Cooperation) IT Equipment and Accessories",  "Surgical and Medical Instrument Manufacturing", "339112", 0.119]
        ]
        
        epa_df = pd.DataFrame(epa_data, columns=['Spend Category', 'NAICS Title', 'NAICS Code', 'kg CO2e/USD'])
        
        # Load spend data
        spend_df = pd.read_excel(
            spend_file,
            usecols=[
                "YearReceipt", "PurchaseOrderNumber", "Supplier", 
                "ShippingMethod", "SpendCategory", "LineDescription",
                "FundType", "ShipToAddressCountry", "Freight per APO", 
                "Amount per PO Line"
            ]
        )
        
        # Clean and standardize spend categories
        spend_df['SpendCategory'] = spend_df['SpendCategory'].str.strip()
        
        # ========== NEW: Extract APO numbers from AWB and BOL data ==========
        apo_numbers_in_transport = set()
        
        # Get APOs from AWB data
        if awb_data is not None and not awb_data.empty:
            apo_numbers_in_transport.update(awb_data['APO'].dropna().astype(str).unique())
        
        # Get APOs from BOL data
        if bol_data is not None and not bol_data.empty:
            apo_numbers_in_transport.update(bol_data['APO no.'].dropna().astype(str).unique())
        
        # Filter spend data to only include APOs that are in transport data
        if apo_numbers_in_transport:
            spend_df = spend_df[spend_df['PurchaseOrderNumber'].astype(str).isin(apo_numbers_in_transport)]
        else:
            st.warning("No APO numbers found in transport data. Scope 3.1 calculation will be empty.")
            return None
        # ========== END NEW ==========
        
        # Apply filters
        filtered_spend = spend_df.copy()
        
        # Year filter
        if selected_years:
            filtered_spend = filtered_spend[
                filtered_spend['YearReceipt'].astype(str).isin(selected_years)
            ]
        
        # FundType filter
        filtered_spend = filtered_spend[
            filtered_spend['FundType'].isin(['Revolving Fund', 'Strategic Fund'])
        ]
        
        # Exclude donations
        filtered_spend = filtered_spend[
            ~filtered_spend['LineDescription'].str.contains('donations|discount', case=False, na=False)
        ]
        
        if filtered_spend.empty:
            st.warning("No spend data matches the selected filters and transport APO boundaries")
            return None
        
        # Merge using exact matching
        merged_df = pd.merge(
            filtered_spend,
            epa_df,
            left_on='SpendCategory',
            right_on='Spend Category',
            how='left'
        )
        
        # Debug: Show matching statistics
        total_rows = len(merged_df)
        matched_rows = len(merged_df[~merged_df['NAICS Code'].isna()])
        #st.info(f"Matching rate: {matched_rows}/{total_rows} ({matched_rows/total_rows:.1%}) rows matched")
        
        if matched_rows == 0:
            st.error("No spend categories matched with EPA factors. Please check your data.")
            return None
        
        # Calculate emissions
        merged_df['kg CO2e'] = merged_df['Amount per PO Line'] * merged_df['kg CO2e/USD']
        merged_df['t CO2e'] = merged_df['kg CO2e'] / 1000
        
        # Final output columns
        result_df = merged_df[[
            'Supplier', 'SpendCategory', 'NAICS Title', 'NAICS Code',
            'Amount per PO Line', 'kg CO2e/USD', 'kg CO2e', 't CO2e',
            'YearReceipt', 'FundType', 'PurchaseOrderNumber'
        ]].rename(columns={
            'SpendCategory': 'Spend Category',
            'Amount per PO Line': 'Amount (USD)'
        })
        
        # Sort and add totals
        result_df = result_df.sort_values('Supplier', ascending=False)
        totals = pd.DataFrame([{
            'Supplier': 'TOTAL',
            'Amount (USD)': result_df['Amount (USD)'].sum(),
            'kg CO2e': result_df['kg CO2e'].sum(),
            't CO2e': result_df['t CO2e'].sum()
        }])
        
        return pd.concat([result_df, totals], ignore_index=True)
        
    except Exception as e:
        st.error(f"Error calculating Scope 3.1 emissions: {e}")
        return None

# ==================== EPA EF Functions ====================
@st.cache_data
def load_epa_ef():
    """Load EPA emission factors from GitHub"""
    try:
        st.info("Loading EPA EF from GitHub...")
        response = requests.get(GITHUB_EPA_EF_URL, timeout=10)
        response.raise_for_status()
        return response.content
    except Exception as e:
        st.error(f"Failed to load EPA EF from GitHub: {str(e)}")
        st.error(f"URL used: {GITHUB_EPA_EF_URL}")
        return None

# ==================== GEO MASTER LOADING ====================
@st.cache_data
def load_geo_master():
    """Load geo_master from GitHub with caching"""
    try:
        st.info("Loading geo_master from GitHub...")
        response = requests.get(GITHUB_GEO_MASTER_URL, timeout=30)
        response.raise_for_status()
        return response.content
    except Exception as e:
        st.error(f"Failed to load geo_master from GitHub: {str(e)}")
        st.error(f"URL used: {GITHUB_GEO_MASTER_URL}")
        return None

# ==================== EXIOBASE SCOPE 3.1 FUNCTION ====================
def calculate_scope31_emissions_exiobase(spend_file, geo_file, awb_data, bol_data, selected_years):
    """Calculate Scope 3.1 emissions using Exiobase emission factors with supplier-country mapping - ONLY for APOs in AWB/BOL"""
    try:
        # Define the two sets of Exiobase emission factors
        # Group 1: Pharmaceutical/medical product categories (ID 24d)
        exiobase_group1 = {
            'AUSTRIA': 0.545131753,
            'BELGIUM': 0.479460088,
            'BRAZIL': 1.109644522,
            'BULGARIA': 1.339848257,
            'CANADA': 0.975906107,
            'CHINA': 2.331383703,
            'CYPRUS': 0.969797487,
            'DENMARK': 0.684560609,
            'FINLAND': 0.494782497,
            'FRANCE': 0.217503994,
            'GERMANY': 0.504585879,
            'GREECE': 1.063976939,
            'INDIA': 2.335286193,
            'INDONESIA': 1.249600634,
            'IRELAND': 0.479919211,
            'ITALY': 0.938398599,
            'NETHERLANDS': 0.976116293,
            'SOUTH KOREA': 1.455342401,
            'SPAIN': 0.753155613,
            'SWEDEN': 0.41669048,
            'SWITZERLAND': 0.641113126,
            'UNITED STATES': 0.515891942,
            'UNITED ARAB EMIRATES': 2.238255645,
            'ISRAEL': 2.238255645,
            'ARGENTINA': 2.605049361,
            'VIETNAM': 2.238255645,
            'PANAMA': 2.605049361,
            'JAPAN': 2.238255645,
            'COSTA RICA': 2.605049361,
            'COLOMBIA': 2.605049361,
            'URUGUAY': 2.605049361,
            'MALAYSIA': 2.238255645,
            'ANTIGUA AND BARBUDA': 2.605049361,
            'THAILAND': 2.238255645,
            'LUXEMBOURG': 0.52,
            'SOUTH AFRICA': 1.90,
            'UNITED KINGDOM': 0.29,
            'MEXICO': 1.09
        }

        # Group 2: Equipment/analyzers categories (ID 33)
        exiobase_group2 = {
            'AUSTRIA': 0.154586475,
            'BELGIUM': 0.440162225,
            'BRAZIL': 0.397086636,
            'BULGARIA': None,
            'CANADA': None,
            'CHINA': 0.980305411,
            'CYPRUS': None,
            'DENMARK': None,
            'FINLAND': None,
            'FRANCE': 0.266389395,
            'GERMANY': 0.180232507,
            'GREECE': None,
            'INDIA': 0.984197791,
            'INDONESIA': None,
            'IRELAND': 0.159060233,
            'ITALY': None,
            'NETHERLANDS': 0.292628813,
            'SOUTH KOREA': 0.474450783,
            'SPAIN': None,
            'SWEDEN': 0.25623846,
            'SWITZERLAND': 0.28,
            'UNITED STATES': 0.387415696,
            'UNITED ARAB EMIRATES': 1.236221092,
            'ISRAEL': 1.236221092,
            'ARGENTINA': 0.662505065,
            'VIETNAM': 1.236221092,
            'PANAMA': 0.662505065,
            'JAPAN': 1.236221092,
            'COSTA RICA': 0.662505065,
            'COLOMBIA': 0.662505065,
            'URUGUAY': 0.662505065,
            'MALAYSIA': 1.236221092,
            'ANTIGUA AND BARBUDA': 0.662505065,
            'THAILAND': 1.236221092,
            'SOUTH AFRICA': 1.02,
            'UNITED KINGDOM': 0.15,
            'MEXICO': 0.8
        }

        # Group 1 categories (ID 24d)
        group1_categories = [
            "(PRO) Antiretroviral medicines",
            "(PRO) Antileishmaniasis",
            "(PRO) Antitubercular drugs",
            "(PRO) Disease vectors management and control",
            "(PRO) Vaccines and antigens and toxoids",
            "(PRO) Antibacterials",
            "(PRO) Anti-malarial",
            "(PRO) Antineoplastic agents",
            "(PRO) Cardiovascular Medicines",
            "(PRO) Nutritional Supplements",
            "(PRO) Antiviral drugs",
            "(PRO) Anti-Chagas",
            "(PRO) Antifungal drugs",
            "(PRO) Syringes and accessories",
            "(PRO) Antipsychotics",
            "(PRO) Central nervous system drugs",
            "(PRO) Needle or blade or sharps disposal container or cart",
            "(PRO) Veterinary vaccines and virology products",
            "(PRO) Hematolic drugs",
            "(PRO) Anaesthetic drugs and related adjuncts and analeptics",
            "(PRO) Patient care and treatment products and supplies",
            "(PRO) Antidotes and emetics",
            "(PRO) Industrial freezers and refrigerators",
            "(PRO) Antidiabetic agents and hyperglycemic agents",
            "(PRO) Cold storage box",
            "(PRO) Cold pack or ice brick",
            "(CC) (Technical Cooperation) Supplies and Materials",
            "(PRO) Drugs affecting the respiratory tract",
            "(PRO) Electrolytes",
            "(PRO) Muscle Relaxant Medicines",
            "(CC) (Technical Cooperation) IT Equipment and Accessories"
        ]

        # Group 2 categories (ID 33)
        group2_categories = [
            "(PRO) Clinical and diagnostic analyzers and accessories and supplies",
            "(PRO) Rapid test kits",
            "(PRO) Laboratory and scientific equipment",
            "(PRO) Laboratory supplies and fixtures",
            "(PRO) Medical Equipment and Accessories and Supplies",
            "(CC) (PAHO internal use) Equipment Services: installation, maintenance, leasing and training",
            "(PRO)Temperature and heat measuring instruments"
        ]

        def standardize_country_name(country_input):
            """Convert any country format to standardized full name"""
            if pd.isna(country_input):
                return None
                
            country_str = str(country_input).upper().strip()
            
            # First check if it's already a standardized name we recognize
            if country_str in exiobase_group1 or country_str in exiobase_group2:
                return country_str
                
            # Try fuzzy matching for known countries
            known_countries = list(exiobase_group1.keys()) + list(exiobase_group2.keys())
            best_match, score = process.extractOne(country_str, known_countries)
            if score > 80:  # Good match threshold
                return best_match
                
            return None  # No match found

        # Load geo master data for supplier-country mapping
        if isinstance(geo_file, str):
            # It's a file path - read it
            with open(geo_file, 'rb') as f:
                geo_content = f.read()
            geo_data = get_geo_master_data(geo_content)
        else:
            # It's already content (BytesIO or bytes)
            geo_data = get_geo_master_data(geo_file)
        
        if geo_data[0] is None:
            st.error("Failed to load geo master data for supplier-country mapping")
            return None
            
        _, _, _, _, supplier_df, _ = geo_data
        
        # Create supplier to country mapping from the supplier worksheet (SAME APPROACH AS SUPPLIER-SPECIFIC METHOD)
        supplier_country_map = {}
        for _, row in supplier_df.iterrows():
            supplier = row['Supplier']
            country = row['country']
            if pd.notna(supplier) and pd.notna(country):
                standardized_country = standardize_country_name(country)
                if standardized_country:
                    supplier_country_map[supplier] = standardized_country
        
        # Load spend data
        spend_df = pd.read_excel(
            spend_file,
            usecols=[
                "YearReceipt", "PurchaseOrderNumber", "Supplier", 
                "ShippingMethod", "SpendCategory", "LineDescription",
                "FundType", "ShipToAddressCountry", "Freight per APO", 
                "Amount per PO Line"
            ]
        )
        
        # Clean and standardize spend categories
        spend_df['SpendCategory'] = spend_df['SpendCategory'].str.strip()
        
        # ========== Extract APO numbers from AWB and BOL data ==========
        apo_numbers_in_transport = set()
        
        # Get APOs from AWB data
        if awb_data is not None and not awb_data.empty:
            apo_numbers_in_transport.update(awb_data['APO'].dropna().astype(str).unique())
        
        # Get APOs from BOL data
        if bol_data is not None and not bol_data.empty:
            apo_numbers_in_transport.update(bol_data['APO no.'].dropna().astype(str).unique())
        
        # Filter spend data to only include APOs that are in transport data
        if apo_numbers_in_transport:
            spend_df = spend_df[spend_df['PurchaseOrderNumber'].astype(str).isin(apo_numbers_in_transport)]
        else:
            st.warning("No APO numbers found in transport data. Scope 3.1 calculation will be empty.")
            return None
        
        # Apply filters
        filtered_spend = spend_df.copy()
        
        # Year filter
        if selected_years:
            filtered_spend = filtered_spend[
                filtered_spend['YearReceipt'].astype(str).isin(selected_years)
            ]
        
        # FundType filter
        filtered_spend = filtered_spend[
            filtered_spend['FundType'].isin(['Revolving Fund', 'Strategic Fund'])
        ]
        
        # Exclude donations
        filtered_spend = filtered_spend[
            ~filtered_spend['LineDescription'].str.contains('donations|discount', case=False, na=False)
        ]
        
        if filtered_spend.empty:
            st.warning("No spend data matches the selected filters and transport APO boundaries")
            return None

        # Add Exiobase calculations to the spend data
        result_df = filtered_spend.copy()
        
        # Determine which group each spend category belongs to and get appropriate EF
        result_df['Exiobase_Group'] = result_df['SpendCategory'].apply(
            lambda x: 'ID 24d' if x in group1_categories else ('ID 33' if x in group2_categories else None)
        )
        
        # Map supplier to standardized country name using the supplier worksheet mapping (SAME APPROACH)
        result_df['Supplier_Country'] = result_df['Supplier'].map(supplier_country_map)
        
        # Get Exiobase emission factor
        def get_exiobase_ef(group, country):
            if pd.isna(group) or pd.isna(country):
                return None
            if group == 'ID 24d':
                return exiobase_group1.get(country, None)
            elif group == 'ID 33':
                return exiobase_group2.get(country, None)
            return None
        
        result_df['Exiobase_kg_CO2e/USD'] = result_df.apply(
            lambda row: get_exiobase_ef(row['Exiobase_Group'], row['Supplier_Country']), axis=1
        )
        
        # Calculate emissions - CORRECTED: Exiobase factors are kg CO2e/USD, not g CO2e/USD
        result_df['Exiobase_kg_CO2e'] = result_df['Amount per PO Line'] * result_df['Exiobase_kg_CO2e/USD']
        result_df['Exiobase_t_CO2e'] = result_df['Exiobase_kg_CO2e'] / 1000  # Convert kg to tons
        
        # Final output columns
        final_columns = [
            'Supplier', 'SpendCategory', 'Supplier_Country', 'Exiobase_Group',
            'Amount per PO Line', 'Exiobase_kg_CO2e/USD', 'Exiobase_kg_CO2e', 'Exiobase_t_CO2e',
            'YearReceipt', 'FundType', 'PurchaseOrderNumber'
        ]
        
        result_df = result_df[final_columns].rename(columns={
            'SpendCategory': 'Spend Category',
            'Amount per PO Line': 'Amount (USD)',
            'Supplier_Country': 'Supplier Country'
        })
        
        # Sort and add totals
        result_df = result_df.sort_values('Supplier', ascending=False)
        totals = pd.DataFrame([{
            'Supplier': 'TOTAL',
            'Amount (USD)': result_df['Amount (USD)'].sum(),
            'Exiobase_kg_CO2e': result_df['Exiobase_kg_CO2e'].sum(),
            'Exiobase_t_CO2e': result_df['Exiobase_t_CO2e'].sum()
        }])
        
        return pd.concat([result_df, totals], ignore_index=True)
        
    except Exception as e:
        st.error(f"Error calculating Scope 3.1 emissions with Exiobase: {e}")
        return None
    
def calculate_scope31_emissions_supplier_specific(spend_file, geo_file, awb_data, bol_data, selected_years):
    """Calculate Scope 3.1 emissions using supplier-specific factors + Exiobase fallback - ONLY for APOs in AWB/BOL"""
    try:
        # Load geo master data for supplier-specific factors and country mapping
        if isinstance(geo_file, str):
            # It's a file path - read it
            with open(geo_file, 'rb') as f:
                geo_content = f.read()
            geo_data = get_geo_master_data(geo_content)
        else:
            # It's already content (BytesIO or bytes)
            geo_data = get_geo_master_data(geo_file)        
        
        if geo_data[0] is None:
            st.error("Failed to load geo master data for supplier-specific calculation")
            return None
        geo_df, _, _, _, supplier_df, precomputed_mappings = geo_data
        
        # Create supplier mapping with both country and supplier-specific EF
        supplier_info_map = {}
        for _, row in supplier_df.iterrows():
            supplier = row['Supplier']
            if pd.notna(supplier):
                supplier_info_map[supplier] = {
                    'country': row['country'] if pd.notna(row['country']) else None,
                    'supplier_specific_ef': row['supplier_specific_ef_sc123'] if pd.notna(row['supplier_specific_ef_sc123']) else None
                }
        
        # Load spend data
        spend_df = pd.read_excel(
            spend_file,
            usecols=[
                "YearReceipt", "PurchaseOrderNumber", "Supplier", 
                "ShippingMethod", "SpendCategory", "LineDescription",
                "FundType", "ShipToAddressCountry", "Freight per APO", 
                "Amount per PO Line"
            ]
        )
        
        # Clean and standardize spend categories
        spend_df['SpendCategory'] = spend_df['SpendCategory'].str.strip()
        
        # ========== NEW: Extract APO numbers from AWB and BOL data ==========
        apo_numbers_in_transport = set()
        
        # Get APOs from AWB data
        if awb_data is not None and not awb_data.empty:
            apo_numbers_in_transport.update(awb_data['APO'].dropna().astype(str).unique())
        
        # Get APOs from BOL data
        if bol_data is not None and not bol_data.empty:
            apo_numbers_in_transport.update(bol_data['APO no.'].dropna().astype(str).unique())
        
        # Filter spend data to only include APOs that are in transport data
        if apo_numbers_in_transport:
            spend_df = spend_df[spend_df['PurchaseOrderNumber'].astype(str).isin(apo_numbers_in_transport)]
        else:
            st.warning("No APO numbers found in transport data. Scope 3.1 calculation will be empty.")
            return None
        # ========== END NEW ==========
        
        # Apply filters
        filtered_spend = spend_df.copy()
        
        # Year filter
        if selected_years:
            filtered_spend = filtered_spend[
                filtered_spend['YearReceipt'].astype(str).isin(selected_years)
            ]
        
        # FundType filter
        filtered_spend = filtered_spend[
            filtered_spend['FundType'].isin(['Revolving Fund', 'Strategic Fund'])
        ]
        
        # Exclude donations
        filtered_spend = filtered_spend[
            ~filtered_spend['LineDescription'].str.contains('donations|discount', case=False, na=False)
        ]
        
        if filtered_spend.empty:
            st.warning("No spend data matches the selected filters and transport APO boundaries")
            return None
        
        # Add supplier information to spend data
        result_df = filtered_spend.copy()
        
        # Extract country and supplier-specific EF directly - FIXED: Handle missing suppliers
        result_df['Supplier_Country'] = result_df['Supplier'].apply(
            lambda x: supplier_info_map.get(x, {}).get('country') if pd.notna(x) else None
        )
        result_df['Supplier_Specific_EF'] = result_df['Supplier'].apply(
            lambda x: supplier_info_map.get(x, {}).get('supplier_specific_ef') if pd.notna(x) else None
        )
        
        # Determine calculation method for each row
        result_df['Calculation_Method'] = result_df['Supplier_Specific_EF'].apply(
            lambda x: 'Supplier-Specific' if pd.notna(x) else 'Exiobase'
        )
        
        # Show method distribution
        #method_counts = result_df['Calculation_Method'].value_counts()
        #st.info(f"Calculation method distribution:\n{method_counts}")
        
        # Get Exiobase factors for fallback (reuse existing function)
        exiobase_result = calculate_scope31_emissions_exiobase(
            spend_file, geo_file, awb_data, bol_data, selected_years
        )
        
        if exiobase_result is None:
            st.error("Failed to get Exiobase factors for fallback calculation")
            return None
        
        # Create a dictionary for Exiobase factors instead of merging
        # This prevents the cartesian product issue with duplicate suppliers
        exiobase_dict = {}
        for _, row in exiobase_result.iterrows():
            if row['Supplier'] != 'TOTAL':  # Skip the total row
                key = (row['Supplier'], row['Spend Category'])
                exiobase_dict[key] = row['Exiobase_kg_CO2e/USD']
        
        # Map Exiobase factors using the dictionary - much more efficient
        result_df['Exiobase_kg_CO2e/USD'] = result_df.apply(
            lambda row: exiobase_dict.get((row['Supplier'], row['SpendCategory']), None),
            axis=1
        )
        
        # Calculate emissions based on method - SIMPLIFIED and SAFE
        def calculate_emissions(row):
            try:
                amount = row['Amount per PO Line']
                if pd.isna(amount) or amount == 0:
                    return 0
                
                if (row['Calculation_Method'] == 'Supplier-Specific' and 
                    pd.notna(row['Supplier_Specific_EF'])):
                    # Use supplier-specific factor
                    ef = row['Supplier_Specific_EF']
                    if ef < 0 or ef > 1000:  # Sanity check: EF should be reasonable
                        st.warning(f"Unusual supplier-specific EF for {row['Supplier']}: {ef}")
                        return 0
                    return ef * amount
                
                elif pd.notna(row['Exiobase_kg_CO2e/USD']):
                    # Use Exiobase factor
                    ef = row['Exiobase_kg_CO2e/USD']
                    if ef < 0 or ef > 1000:  # Sanity check
                        st.warning(f"Unusual Exiobase EF for {row['Supplier']}: {ef}")
                        return 0
                    return ef * amount
                
                else:
                    return 0  # No valid factors
                    
            except Exception as e:
                # Silent error handling to avoid breaking the entire calculation
                return 0
        
        result_df['kg_CO2e'] = result_df.apply(calculate_emissions, axis=1)
        result_df['t_CO2e'] = result_df['kg_CO2e'] / 1000
        
        # Check for astronomical values
        total_emissions = result_df['kg_CO2e'].sum()
        max_emission = result_df['kg_CO2e'].max()
        
        #st.info(f"Total emissions: {total_emissions:,.2f} kg CO2e")
        #st.info(f"Maximum row emission: {max_emission:,.2f} kg CO2e")
        
        if total_emissions > 1e9:  # More than 1 billion kg CO2e
            st.warning("⚠️ Very large total emissions detected. Checking data...")
            
            # Check supplier-specific factors
            supplier_specific_ef_stats = result_df[result_df['Calculation_Method'] == 'Supplier-Specific']['Supplier_Specific_EF'].describe()
            st.warning(f"Supplier-specific EF stats:\n{supplier_specific_ef_stats}")
            
            # Check spend amounts
            amount_stats = result_df['Amount per PO Line'].describe()
            st.warning(f"Spend amount stats:\n{amount_stats}")
            
            # Show top 10 emitting rows
            top_emitters = result_df.nlargest(10, 'kg_CO2e')[['Supplier', 'Amount per PO Line', 'Supplier_Specific_EF', 'Exiobase_kg_CO2e/USD', 'kg_CO2e']]
            st.warning(f"Top 10 emitting rows:\n{top_emitters}")
        
        # Final output columns
        final_columns = [
            'Supplier', 'SpendCategory', 'Supplier_Country', 'Calculation_Method',
            'Supplier_Specific_EF', 'Exiobase_kg_CO2e/USD', 'Amount per PO Line', 
            'kg_CO2e', 't_CO2e', 'YearReceipt', 'FundType', 'PurchaseOrderNumber'
        ]
        
        result_df = result_df[final_columns].rename(columns={
            'SpendCategory': 'Spend Category',
            'Amount per PO Line': 'Amount (USD)',
            'Supplier_Country': 'Supplier Country',
            'Supplier_Specific_EF': 'Supplier Specific EF (kg CO2e/USD)',
            'Exiobase_kg_CO2e/USD': 'Exiobase EF (kg CO2e/USD)'
        })
        
        # Sort and add totals
        result_df = result_df.sort_values('Supplier', ascending=False)
        
        # Create totals row without including it in the main dataframe
        totals = pd.DataFrame([{
            'Supplier': 'TOTAL',
            'Spend Category': '',
            'Supplier Country': '',
            'Calculation_Method': '',
            'Supplier Specific EF (kg CO2e/USD)': '',
            'Exiobase EF (kg CO2e/USD)': '',
            'Amount (USD)': result_df['Amount (USD)'].sum(),
            'kg_CO2e': result_df['kg_CO2e'].sum(),
            't_CO2e': result_df['t_CO2e'].sum(),
            'YearReceipt': '',
            'FundType': '',
            'PurchaseOrderNumber': ''
        }])
        
        return pd.concat([result_df, totals], ignore_index=True)
        
    except Exception as e:
        st.error(f"Error calculating Scope 3.1 emissions with Supplier-Specific + Exiobase: {e}")
        import traceback
        st.error(f"Traceback: {traceback.format_exc()}")
        return None

def get_top_flight_routes(awb_data, n=5):
    """Get top flight routes by count (number of flights)"""
    if awb_data is None or awb_data.empty:
        return pd.DataFrame()
    
    # Filter out TOTAL rows and get only flight legs
    flight_data = awb_data[awb_data['leg'] != 'TOTAL'].copy()
    
    if flight_data.empty:
        return pd.DataFrame()
    
    # Group by origin-destination pairs and count flights
    route_counts = flight_data.groupby(['origin', 'destination']).agg({
        'leg': 'count',
        'ghg_emissions_tCO2e': 'sum',
        'distance_km': 'mean'
    }).rename(columns={
        'leg': 'flight_count',
        'distance_km': 'avg_distance_km'
    }).reset_index()
    
    # Create route label
    route_counts['route'] = route_counts['origin'] + ' → ' + route_counts['destination']
    
    # Sort by flight count and get top N
    top_routes = route_counts.nlargest(n, 'flight_count')
    
    return top_routes[['route', 'flight_count', 'ghg_emissions_tCO2e', 'avg_distance_km']]

def get_top_source_locations(awb_data, n=5):
    """Get top true source locations by count (number of shipments) - only first origins from AWB data only"""
    source_counts = []
    
    # Process AWB data only - only get the FIRST origin for each AWB
    if awb_data is not None and not awb_data.empty:
        # Get the first leg (lowest leg number) for each AWB to find the true origin
        first_origins = awb_data[awb_data['leg'] != 'TOTAL'].copy()
        first_origins = first_origins.sort_values(['awb_row', 'leg']).groupby('awb_row').first().reset_index()
        
        awb_sources = first_origins.groupby('origin').agg({
            'awb_row': 'count',
            'ghg_emissions_tCO2e': 'sum'
        }).rename(columns={
            'awb_row': 'shipment_count'
        }).reset_index()
        awb_sources['transport_type'] = 'Air'
        source_counts.append(awb_sources.rename(columns={'origin': 'location'}))
    
    if not source_counts:
        return pd.DataFrame()
    
    # Combine and get top locations by shipment count
    all_sources = pd.concat(source_counts, ignore_index=True)
    top_locations = all_sources.groupby('location').agg({
        'shipment_count': 'sum',
        'ghg_emissions_tCO2e': 'sum'
    }).reset_index()
    
    # Get transport type composition for each location
    transport_composition = all_sources.groupby(['location', 'transport_type'])['shipment_count'].sum().unstack(fill_value=0)
    transport_composition = transport_composition.reset_index()
    
    # Merge composition data
    top_locations = top_locations.merge(transport_composition, on='location', how='left')
    
    # Sort by shipment count and get top N
    top_locations = top_locations.nlargest(n, 'shipment_count')
    
    return top_locations

# ==================== TCE 1 CALCULATION FUNCTIONS ====================
def get_continent_from_coords(lat, lon):
    """Get continent from latitude and longitude coordinates"""
    if pd.isna(lat) or pd.isna(lon):
        return None
    
    # Simple continent mapping based on coordinates
    # This is a simplified approach - you might want to use a more robust geocoding service
    if -170 <= lon <= -30:  # Americas
        if 10 <= lat <= 80:  # North America
            return "North America"
        elif -60 <= lat <= 15:  # South America
            return "South America"
    elif -20 <= lon <= 60:  # Africa/Europe
        if -40 <= lat <= 40:  # Africa
            return "Africa"
        elif 35 <= lat <= 75:  # Europe
            return "Europe"
    elif 60 <= lon <= 180:  # Asia/Australia
        if -15 <= lat <= 60:  # Asia
            return "Asia"
    
    return None

def get_distance_by_continent(continent):
    """Get distance based on continent using the provided values"""
    distance_map = {
        "Asia": 662.00,
        "Europe": 100.00,
        "North America": 330.00,
        "South America": 600.00,
        "Africa": 618.31
    }
    return distance_map.get(continent, 0.0)

def calculate_tce1_emissions(weight_kg, distance_km):
    """Calculate TCE 1 emissions using the provided emission factor"""
    if pd.isna(weight_kg) or weight_kg <= 0 or pd.isna(distance_km) or distance_km <= 0:
        return 0.0
    
    # Convert kg to tons
    weight_ton = weight_kg / 1000
    
    # Emission factor: 861.4 g CO2e/ton-km
    ef = 861.4
    
    # Calculate emissions in g CO2e
    emissions_g = weight_ton * distance_km * ef
    
    # Convert to t CO2e
    emissions_t = emissions_g / 1000000
    
    return emissions_t

def calculate_tce1_breakdown(awb_data, bol_data, precomputed_mappings):
    """Calculate TCE 1 emissions for both air and sea shipments"""
    tce1_results = {'Air': 0.0, 'Sea': 0.0, 'Total': 0.0}
    
    # Process AWB data (Air shipments)
    if awb_data is not None and not awb_data.empty:
        # Get only the first leg of each AWB to get the true origin
        first_legs = awb_data[awb_data['leg'] != 'TOTAL'].copy()
        first_legs = first_legs.sort_values(['awb_row', 'leg']).groupby('awb_row').first().reset_index()
        
        for _, row in first_legs.iterrows():
            if pd.notna(row['gross_weight']) and row['gross_weight'] > 0:
                # Get continent from coordinates
                continent = None
                if pd.notna(row['origin_lat']) and pd.notna(row['origin_lon']):
                    continent = get_continent_from_coords(row['origin_lat'], row['origin_lon'])
                
                # If continent not found from coordinates, try to get from airport code
                if continent is None and pd.notna(row['origin']):
                    try:
                        origin_coords = get_airport_coords(row['origin'])
                        if None not in origin_coords:
                            continent = get_continent_from_coords(origin_coords[0], origin_coords[1])
                    except:
                        pass
                
                # Get distance based on continent
                distance_km = get_distance_by_continent(continent) if continent else 0.0
                
                # Calculate TCE 1 emissions
                tce1_emissions = calculate_tce1_emissions(row['gross_weight'], distance_km)
                
                tce1_results['Air'] += tce1_emissions
                tce1_results['Total'] += tce1_emissions
    
    # Process BOL data (Sea shipments)
    if bol_data is not None and not bol_data.empty:
        # Use precomputed city_to_continent mapping
        city_to_continent = precomputed_mappings['city_to_continent']
        
        for _, row in bol_data.iterrows():
            if pd.notna(row['Gross weight, kg']) and row['Gross weight, kg'] > 0:
                # Get continent from BOL data using precomputed mapping
                continent = None
                if 'Loading Port Continent' in row and pd.notna(row['Loading Port Continent']):
                    continent = row['Loading Port Continent']
                elif pd.notna(row['Port of loading']):
                    port_upper = row['Port of loading'].upper()
                    continent = city_to_continent.get(port_upper)
                
                # Get distance based on continent
                distance_km = get_distance_by_continent(continent) if continent else 0.0
                
                # Calculate TCE 1 emissions
                tce1_emissions = calculate_tce1_emissions(row['Gross weight, kg'], distance_km)
                
                tce1_results['Sea'] += tce1_emissions
                tce1_results['Total'] += tce1_emissions
    
    return tce1_results

# ==================== TCE 2 & 4 CALCULATION FUNCTIONS ====================
def calculate_tce_emissions(weight_kg, transport_mode, cargo_type):
    """
    Calculate TCE 2 or TCE 4 emissions based on transport mode and cargo type
    
    Parameters:
    - weight_kg: Weight in kilograms
    - transport_mode: 'Air' or 'Sea'
    - cargo_type: 'Reefer' or 'Dry'
    """
    if pd.isna(weight_kg) or weight_kg <= 0:
        return 0.0
    
    # Convert kg to tons
    weight_ton = weight_kg / 1000
    
    # Define emission factors (g CO2e/ton)
    if transport_mode == 'Air':
        if cargo_type == 'Reefer':
            ef = 2500  # g CO2e/ton for air reefer
        else:
            ef = 1300  # g CO2e/ton for air dry
    else:  # Sea
        if cargo_type == 'Reefer':
            ef = 1260  # g CO2e/ton for sea reefer
        else:
            ef = 1070  # g CO2e/ton for sea dry
    
    # Calculate emissions in g CO2e
    emissions_g = weight_ton * ef
    
    # Convert to t CO2e
    emissions_t = emissions_g / 1000000
    
    return emissions_t

def calculate_tce2_breakdown(awb_data, bol_data):
    """Calculate TCE 2 emissions (export hub operations) for both air and sea shipments"""
    tce2_results = {'Air': 0.0, 'Sea': 0.0, 'Total': 0.0}
    
    # Process AWB data (Air shipments)
    if awb_data is not None and not awb_data.empty:
        # Get only the first leg of each AWB to get the true origin data
        first_legs = awb_data[awb_data['leg'] != 'TOTAL'].copy()
        first_legs = first_legs.sort_values(['awb_row', 'leg']).groupby('awb_row').first().reset_index()
        
        for _, row in first_legs.iterrows():
            if pd.notna(row['gross_weight']) and row['gross_weight'] > 0:
                # Determine cargo type based on temperature control
                cargo_type = 'Reefer' if row.get('Temperature Control') == 'YES' else 'Dry'
                
                # Calculate TCE 2 emissions
                tce2_emissions = calculate_tce_emissions(row['gross_weight'], 'Air', cargo_type)
                
                tce2_results['Air'] += tce2_emissions
                tce2_results['Total'] += tce2_emissions
    
    # Process BOL data (Sea shipments)
    if bol_data is not None and not bol_data.empty:
        for _, row in bol_data.iterrows():
            if pd.notna(row['Gross weight, kg']) and row['Gross weight, kg'] > 0:
                # Determine cargo type based on temperature control
                cargo_type = 'Reefer' if row.get('Temperature Control') == 'YES' else 'Dry'
                
                # Calculate TCE 2 emissions
                tce2_emissions = calculate_tce_emissions(row['Gross weight, kg'], 'Sea', cargo_type)
                
                tce2_results['Sea'] += tce2_emissions
                tce2_results['Total'] += tce2_emissions
    
    return tce2_results

def calculate_tce4_breakdown(awb_data, bol_data):
    """Calculate TCE 4 emissions (import hub operations) for both air and sea shipments"""
    tce4_results = {'Air': 0.0, 'Sea': 0.0, 'Total': 0.0}
    
    # Process AWB data (Air shipments)
    if awb_data is not None and not awb_data.empty:
        # Get only the first leg of each AWB to get the true origin data
        first_legs = awb_data[awb_data['leg'] != 'TOTAL'].copy()
        first_legs = first_legs.sort_values(['awb_row', 'leg']).groupby('awb_row').first().reset_index()
        
        for _, row in first_legs.iterrows():
            if pd.notna(row['gross_weight']) and row['gross_weight'] > 0:
                # Determine cargo type based on temperature control
                cargo_type = 'Reefer' if row.get('Temperature Control') == 'YES' else 'Dry'
                
                # Calculate TCE 4 emissions
                tce4_emissions = calculate_tce_emissions(row['gross_weight'], 'Air', cargo_type)
                
                tce4_results['Air'] += tce4_emissions
                tce4_results['Total'] += tce4_emissions
    
    # Process BOL data (Sea shipments)
    if bol_data is not None and not bol_data.empty:
        for _, row in bol_data.iterrows():
            if pd.notna(row['Gross weight, kg']) and row['Gross weight, kg'] > 0:
                # Determine cargo type based on temperature control
                cargo_type = 'Reefer' if row.get('Temperature Control') == 'YES' else 'Dry'
                
                # Calculate TCE 4 emissions
                tce4_emissions = calculate_tce_emissions(row['Gross weight, kg'], 'Sea', cargo_type)
                
                tce4_results['Sea'] += tce4_emissions
                tce4_results['Total'] += tce4_emissions
    
    return tce4_results

# ==================== INTENSITY CALCULATION FUNCTIONS ====================
def calculate_scope31_intensity(scope31_results, spend_file, selected_years, method_type="epa"):
    """Calculate Scope 3.1 intensity (g CO2e/$) using FILTERED spend data."""
    if scope31_results is None or scope31_results.empty:
        return None

    # Get total emissions in grams based on method type
    total_emissions_g = 0
    if method_type == "epa":
        if 'kg CO2e' in scope31_results.columns:
            total_row = scope31_results[scope31_results['Supplier'] == 'TOTAL']
            if not total_row.empty:
                total_kg = total_row['kg CO2e'].values[0]
                total_emissions_g = total_kg * 1000  # Convert kg to grams
    elif method_type == "exiobase":
        if 'Exiobase_kg_CO2e' in scope31_results.columns:
            total_row = scope31_results[scope31_results['Supplier'] == 'TOTAL']
            if not total_row.empty:
                total_kg = total_row['Exiobase_kg_CO2e'].values[0]
                total_emissions_g = total_kg * 1000
    elif method_type == "supplier":
        if 'kg_CO2e' in scope31_results.columns:
            total_row = scope31_results[scope31_results['Supplier'] == 'TOTAL']
            if not total_row.empty:
                total_kg = total_row['kg_CO2e'].values[0]
                total_emissions_g = total_kg * 1000

    # Load spend data with FILTERS applied
    try:
        spend_df = pd.read_excel(
            spend_file,
            usecols=["YearReceipt", "FundType", "LineDescription", "Amount per PO Line"]
        )
        
        # Apply filters (same logic as Scope 3.1 calculation)
        filtered_spend = spend_df.copy()
        
        # Year filter
        if selected_years:
            filtered_spend = filtered_spend[
                filtered_spend['YearReceipt'].astype(str).isin(selected_years)
            ]
        
        # FundType filter (Strategic/Revolving only)
        filtered_spend = filtered_spend[
            filtered_spend['FundType'].isin(['Revolving Fund', 'Strategic Fund'])
        ]
        
        # Exclude donations/discounts (case-insensitive)
        filtered_spend = filtered_spend[
            ~filtered_spend['LineDescription'].str.contains('donations|discount', case=False, na=False)
        ]
        
        total_spend = filtered_spend['Amount per PO Line'].sum()
        
        if total_spend > 0 and total_emissions_g > 0:
            return total_emissions_g / total_spend
        else:
            return None
    except Exception as e:
        print(f"Error calculating filtered Scope 3.1 intensity: {e}")
        return None

def calculate_scope34_intensity(awb_data, bol_data):
    """Calculate Scope 3.4 intensity (g CO2e/ton-km)"""
    total_emissions_g = 0
    total_ton_km = 0
    
    # Process AWB data
    if awb_data is not None and not awb_data.empty:
        # Filter out TOTAL rows
        awb_legs = awb_data[awb_data['leg'] != 'TOTAL']
        
        # Sum emissions (convert from tCO2e to gCO2e)
        if 'ghg_emissions_tCO2e' in awb_legs.columns:
            total_emissions_g += awb_legs['ghg_emissions_tCO2e'].sum() * 1000000
        
        # Calculate ton-km for air shipments
        for _, row in awb_legs.iterrows():
            if (pd.notna(row['gross_weight_ton']) and row['gross_weight_ton'] > 0 and 
                pd.notna(row['distance_km']) and row['distance_km'] > 0):
                total_ton_km += row['gross_weight_ton'] * row['distance_km']
    
    # Process BOL data
    if bol_data is not None and not bol_data.empty:
        # Sum emissions (convert from tCO2e to gCO2e)
        if 'ghg_emissions_tCO2e' in bol_data.columns:
            total_emissions_g += bol_data['ghg_emissions_tCO2e'].sum() * 1000000
        
        # Calculate ton-km for ocean shipments
        for _, row in bol_data.iterrows():
            if (pd.notna(row['Gross weight, kg']) and row['Gross weight, kg'] > 0 and 
                pd.notna(row['Sea Distance (km)']) and row['Sea Distance (km)'] > 0):
                weight_ton = row['Gross weight, kg'] / 1000  # Convert kg to tons
                total_ton_km += weight_ton * row['Sea Distance (km)']
    
    # Calculate intensity
    if total_ton_km > 0:
        intensity = total_emissions_g / total_ton_km
        return intensity
    else:
        return None


# ==================== UPDATED MAIN FUNCTION ====================
def main():
    st.title("✈️ GHG Emissions Calculator")
    st.markdown("Calculate emissions from Air Waybills (AWB) and Bills of Lading (BOL)")

    # ==================== SIDEBAR FILTERS ====================
    st.sidebar.header("Filters")
    
    # Analysis scope selection - FIXED KEY
    analysis_scope = st.sidebar.radio(
        "Select Scope to Analyze:",
        options=["Scope 3.1 (Purchased Goods)", "Scope 3.4 (Transportation)", "Both"],
        index=2,
        key="main_scope_selector"
    )
    
    # Single year selection - FIXED KEY
    selected_year = st.sidebar.selectbox(
        "Select Year",
        options=["2024", "2023", "2022"],
        index=0,
        key="year_filter_selectbox"
    )

    # Initialize empty lists for supplier and APO filters
    selected_suppliers = []
    selected_apos = []

    # ==================== FILE UPLOADS ====================
    with st.expander("📁 Upload Required Files", expanded=True):
        col1, col2 = st.columns(2)
        
        with col1:
            main_file = st.file_uploader(
                "1. Main Data (with AWB/BoL worksheets)",
                type=['xlsx'],
                key="main_data_uploader"
            )
            
        with col2:
            spend_file = st.file_uploader(
                "2. Spend Data (PAHO Spend DATA 22-24.xlsx)",
                type=['xlsx'],
                key="spend_data_uploader"
            )

    if not main_file:
        st.info("Please upload the main data file to begin")
        st.stop()

    if analysis_scope in ["Scope 3.1 (Purchased Goods)", "Both"] and not spend_file:
        st.error("Please upload the spend data file for Scope 3.1 analysis")
        st.stop()

    # ==================== GEO MASTER LOADING ====================
    geo_content = load_geo_master()
    if geo_content is None:
        st.stop()

    # Load geo master with precomputed mappings once
    geo_data = get_geo_master_data(geo_content)
    if geo_data[0] is None:
        st.error("Failed to load geo master data")
        st.stop()

    geo_df, cerdi_df, ef_df, reefer_df, supplier_df, precomputed_mappings = geo_data

    # ==================== EPA EF LOADING ====================
    if analysis_scope in ["Scope 3.1 (Purchased Goods)", "Both"]:
        epa_content = load_epa_ef()
        if epa_content is None:
            st.error("EPA emission factors are required for Scope 3.1 calculation")
            st.stop()
    else:
        epa_content = None

    # ==================== FILE PROCESSING ====================
    try:
        xls = pd.ExcelFile(main_file)
        sheet_names = xls.sheet_names
        has_awb = 'AWB' in sheet_names
        has_bol = 'BoL' in sheet_names
        
        if not has_awb and not has_bol:
            st.error("File must contain either 'AWB' or 'BoL' worksheet")
            st.stop()
    except Exception as e:
        st.error(f"Error reading file: {str(e)}")
        st.stop()

    # Initialize variables to avoid UnboundLocalError
    bol_with_tce = None
    awb_with_tce = None

    # Process Scope 3.4 data first
    awb_results = None
    bol_results = None

    if has_awb:
        with st.spinner("Processing AWB data..."):
            awb_results = process_awb_file(main_file, sheet_name='AWB', 
                                        spend_file=spend_file, geo_content=geo_content)

    if has_bol and spend_file:
        with st.spinner("Processing BOL data..."):
            bol_results = process_bol_file(main_file, geo_content, spend_file)
            
            # Create bol_with_tce from bol_results
            if bol_results is not None:
                bol_with_tce = bol_results.copy()
                
                # ===== ADD TCE COLUMNS AND CALCULATIONS HERE =====
                # Initialize TCE columns
                bol_with_tce['TCE1_Emissions_tCO2e'] = 0.0
                bol_with_tce['TCE2_Emissions_tCO2e'] = 0.0
                bol_with_tce['TCE4_Emissions_tCO2e'] = 0.0
                
                # Use precomputed city_to_continent mapping instead of loading geo data again
                city_to_continent = precomputed_mappings['city_to_continent']
                
                # Calculate TCE emissions for each BOL row
                for idx, row in bol_with_tce.iterrows():
                    if pd.notna(row['Gross weight, kg']) and row['Gross weight, kg'] > 0:
                        # TCE 1: Inland transport to export hub
                        continent = None
                        if 'Loading Port Continent' in row and pd.notna(row['Loading Port Continent']):
                            continent = row['Loading Port Continent']
                        elif pd.notna(row['Port of loading']):
                            port_upper = row['Port of loading'].upper()
                            continent = city_to_continent.get(port_upper)
                            
                        distance_km = get_distance_by_continent(continent) if continent else 0.0
                        tce1_emissions = calculate_tce1_emissions(row['Gross weight, kg'], distance_km)
                        bol_with_tce.at[idx, 'TCE1_Emissions_tCO2e'] = tce1_emissions
                        
                        # TCE 2: Export hub operations
                        cargo_type = 'Reefer' if row.get('Temperature Control') == 'YES' else 'Dry'
                        tce2_emissions = calculate_tce_emissions(row['Gross weight, kg'], 'Sea', cargo_type)
                        bol_with_tce.at[idx, 'TCE2_Emissions_tCO2e'] = tce2_emissions
                        
                        # TCE 4: Import hub operations (same calculation as TCE 2 for simplicity)
                        tce4_emissions = calculate_tce_emissions(row['Gross weight, kg'], 'Sea', cargo_type)
                        bol_with_tce.at[idx, 'TCE4_Emissions_tCO2e'] = tce4_emissions
            else:
                bol_with_tce = None

    # ==================== UPDATE FILTERS ====================
    if awb_results is not None or bol_results is not None:
        # Get unique suppliers and APOs from both datasets
        all_suppliers = set()
        all_apos = set()
        
        if awb_results is not None and not awb_results.empty:
            all_suppliers.update(awb_results['Supplier'].dropna().unique())
            all_apos.update(awb_results['APO'].dropna().unique())
            
        if bol_results is not None and not bol_results.empty:
            all_suppliers.update(bol_results['Shipper name'].dropna().unique())
            all_apos.update(bol_results['APO no.'].dropna().unique())
        
        # Convert to sorted lists for the select boxes
        supplier_list = sorted([s for s in all_suppliers if pd.notna(s)])
        apo_list = sorted([str(a) for a in all_apos if pd.notna(a)])
        
        # Create sidebar filters
        selected_suppliers = st.sidebar.multiselect(
            "Filter by Supplier",
            options=supplier_list,
            default=[],
            key="supplier_filter"
        )
        
        selected_apos = st.sidebar.multiselect(
            "Filter by APO no.",
            options=apo_list,
            default=[],
            key="apo_filter"
        )

    # ==================== FILTER DATA ====================
    def filter_data(df, is_awb=True):
        if df is None or df.empty:
            return df
            
        filtered = df.copy()
        
        # Apply supplier filter
        if selected_suppliers:
            if is_awb:
                filtered = filtered[filtered['Supplier'].isin(selected_suppliers)]
            else:
                filtered = filtered[filtered['Shipper name'].isin(selected_suppliers)]
        
        # Apply APO filter
        if selected_apos:
            if is_awb:
                filtered = filtered[filtered['APO'].astype(str).isin(selected_apos)]
            else:
                filtered = filtered[filtered['APO no.'].astype(str).isin(selected_apos)]
        
        return filtered

    filtered_awb = filter_data(awb_results, is_awb=True)
    filtered_bol = filter_data(bol_results, is_awb=False)

    # ==================== SCOPE 3.1 PROCESSING ====================
    # NOW process Scope 3.1 data AFTER filtering
    scope31_results = None
    scope31_exiobase_results = None
    scope31_supplier_specific_results = None

    if analysis_scope in ["Scope 3.1 (Purchased Goods)", "Both"] and spend_file and epa_content:
        with st.spinner("Processing Scope 3.1 data..."):
            scope31_results = calculate_scope31_emissions(spend_file, filtered_awb, filtered_bol, [selected_year])
            scope31_exiobase_results = calculate_scope31_emissions_exiobase(
                spend_file, geo_content, filtered_awb, filtered_bol, [selected_year]
            )
            scope31_supplier_specific_results = calculate_scope31_emissions_supplier_specific(
                spend_file, geo_content, filtered_awb, filtered_bol, [selected_year]
            )

    # ==================== EMISSIONS SUMMARY ====================
    st.header("Emissions Summary")

    # First calculate the basic AWB and BOL emissions (needed for TCE 3)
    awb_emissions = 0.0
    if filtered_awb is not None and not filtered_awb.empty and 'ghg_emissions_tCO2e' in filtered_awb.columns:
        awb_emissions = filtered_awb.loc[filtered_awb['leg'] != 'TOTAL', 'ghg_emissions_tCO2e'].sum()

    bol_emissions = 0.0
    if filtered_bol is not None and not filtered_bol.empty and 'ghg_emissions_tCO2e' in filtered_bol.columns:
        bol_emissions = filtered_bol['ghg_emissions_tCO2e'].sum()

    # Calculate TCE emissions
    tce1_results = calculate_tce1_breakdown(filtered_awb, filtered_bol, precomputed_mappings)
    tce2_results = calculate_tce2_breakdown(filtered_awb, filtered_bol)
    tce3_value = awb_emissions + bol_emissions
    tce4_results = calculate_tce4_breakdown(filtered_awb, filtered_bol)

    # ==================== ADD TCE COLUMNS TO AWB DATA ====================
    if filtered_awb is not None and not filtered_awb.empty:
        # Create a copy to avoid modifying the original
        awb_with_tce = filtered_awb.copy()
        
        # Add TCE columns - for AWB, we'll add these to the first leg of each shipment
        # Get first legs only (to avoid duplicating TCE 1,2,4 across multiple legs)
        first_leg_mask = (awb_with_tce['leg'] != 'TOTAL') & ~awb_with_tce.duplicated('awb_row', keep='first')
        total_mask = awb_with_tce['leg'] == 'TOTAL'
        
        # Initialize TCE columns
        awb_with_tce['TCE1_Emissions_tCO2e'] = 0.0
        awb_with_tce['TCE2_Emissions_tCO2e'] = 0.0  
        awb_with_tce['TCE4_Emissions_tCO2e'] = 0.0
        
        # Calculate TCE emissions for each first leg
        for idx, row in awb_with_tce[first_leg_mask].iterrows():
            if pd.notna(row['gross_weight']) and row['gross_weight'] > 0:
                # TCE 1: Inland transport to export hub
                continent = None
                if pd.notna(row['origin_lat']) and pd.notna(row['origin_lon']):
                    continent = get_continent_from_coords(row['origin_lat'], row['origin_lon'])
                elif pd.notna(row['origin']):
                    try:
                        origin_coords = get_airport_coords(row['origin'])
                        if None not in origin_coords:
                            continent = get_continent_from_coords(origin_coords[0], origin_coords[1])
                    except:
                        pass
                
                distance_km = get_distance_by_continent(continent) if continent else 0.0
                tce1_emissions = calculate_tce1_emissions(row['gross_weight'], distance_km)
                awb_with_tce.at[idx, 'TCE1_Emissions_tCO2e'] = tce1_emissions
                
                # TCE 2: Export hub operations
                cargo_type = 'Reefer' if row.get('Temperature Control') == 'YES' else 'Dry'
                tce2_emissions = calculate_tce_emissions(row['gross_weight'], 'Air', cargo_type)
                awb_with_tce.at[idx, 'TCE2_Emissions_tCO2e'] = tce2_emissions
                
                # TCE 4: Import hub operations (same calculation as TCE 2 for simplicity)
                tce4_emissions = calculate_tce_emissions(row['gross_weight'], 'Air', cargo_type)
                awb_with_tce.at[idx, 'TCE4_Emissions_tCO2e'] = tce4_emissions
        
        # For TOTAL rows, sum up the TCE emissions from all legs
        for apo in awb_with_tce[total_mask]['APO'].unique():
            apo_mask = (awb_with_tce['APO'] == apo) & (awb_with_tce['leg'] != 'TOTAL')
            tce1_total = awb_with_tce.loc[apo_mask, 'TCE1_Emissions_tCO2e'].sum()
            tce2_total = awb_with_tce.loc[apo_mask, 'TCE2_Emissions_tCO2e'].sum()
            tce4_total = awb_with_tce.loc[apo_mask, 'TCE4_Emissions_tCO2e'].sum()
            
            total_row_mask = (awb_with_tce['APO'] == apo) & (awb_with_tce['leg'] == 'TOTAL')
            awb_with_tce.loc[total_row_mask, 'TCE1_Emissions_tCO2e'] = tce1_total
            awb_with_tce.loc[total_row_mask, 'TCE2_Emissions_tCO2e'] = tce2_total
            awb_with_tce.loc[total_row_mask, 'TCE4_Emissions_tCO2e'] = tce4_total
    else:
        # Handle case when there's no AWB data
        awb_with_tce = None

    # ==================== ADD TCE COLUMNS TO BOL DATA ====================
    if filtered_bol is not None and not filtered_bol.empty:
        # Create a copy to avoid modifying the original
        bol_with_tce = filtered_bol.copy()
        
        # Initialize TCE columns
        bol_with_tce['TCE1_Emissions_tCO2e'] = 0.0
        bol_with_tce['TCE2_Emissions_tCO2e'] = 0.0
        bol_with_tce['TCE4_Emissions_tCO2e'] = 0.0
        
        # Use precomputed city_to_continent mapping instead of loading geo data again
        city_to_continent = precomputed_mappings['city_to_continent']
        
        # Calculate TCE emissions for each BOL row
        for idx, row in bol_with_tce.iterrows():
            if pd.notna(row['Gross weight, kg']) and row['Gross weight, kg'] > 0:
                # TCE 1: Inland transport to export hub
                continent = None
                if 'Loading Port Continent' in row and pd.notna(row['Loading Port Continent']):
                    continent = row['Loading Port Continent']
                elif pd.notna(row['Port of loading']):
                    port_upper = row['Port of loading'].upper()
                    continent = city_to_continent.get(port_upper)
                
                distance_km = get_distance_by_continent(continent) if continent else 0.0
                tce1_emissions = calculate_tce1_emissions(row['Gross weight, kg'], distance_km)
                bol_with_tce.at[idx, 'TCE1_Emissions_tCO2e'] = tce1_emissions
                
                # TCE 2: Export hub operations
                cargo_type = 'Reefer' if row.get('Temperature Control') == 'YES' else 'Dry'
                tce2_emissions = calculate_tce_emissions(row['Gross weight, kg'], 'Sea', cargo_type)
                bol_with_tce.at[idx, 'TCE2_Emissions_tCO2e'] = tce2_emissions
                
                # TCE 4: Import hub operations (same calculation as TCE 2 for simplicity)
                tce4_emissions = calculate_tce_emissions(row['Gross weight, kg'], 'Sea', cargo_type)
                bol_with_tce.at[idx, 'TCE4_Emissions_tCO2e'] = tce4_emissions

    # Scope 3.1 totals (with proper error handling)
    scope31_total = 0.0
    if scope31_results is not None and not scope31_results.empty and 'Supplier' in scope31_results.columns:
        total_row = scope31_results[scope31_results['Supplier'] == 'TOTAL']
        if not total_row.empty and 't CO2e' in total_row.columns:
            scope31_total = total_row['t CO2e'].values[0]

    scope31_exiobase_total = 0.0
    if scope31_exiobase_results is not None and not scope31_exiobase_results.empty and 'Supplier' in scope31_exiobase_results.columns:
        total_row = scope31_exiobase_results[scope31_exiobase_results['Supplier'] == 'TOTAL']
        if not total_row.empty and 'Exiobase_t_CO2e' in total_row.columns:
            scope31_exiobase_total = total_row['Exiobase_t_CO2e'].values[0]

    scope31_supplier_total = 0.0
    if scope31_supplier_specific_results is not None and not scope31_supplier_specific_results.empty and 'Supplier' in scope31_supplier_specific_results.columns:
        total_row = scope31_supplier_specific_results[scope31_supplier_specific_results['Supplier'] == 'TOTAL']
        if not total_row.empty and 't_CO2e' in total_row.columns:
            scope31_supplier_total = total_row['t_CO2e'].values[0]

    # Scope 3.1 section - Single header with reordered columns
    st.markdown("#### Scope 3.1 Emissions - Purchased Goods and Services")

    # Calculate intensities for each method
    epa_intensity = calculate_scope31_intensity(scope31_results, spend_file, [selected_year], "epa") if scope31_results is not None else None
    exiobase_intensity = calculate_scope31_intensity(scope31_exiobase_results, spend_file, [selected_year], "exiobase") if scope31_exiobase_results is not None else None
    supplier_intensity = calculate_scope31_intensity(scope31_supplier_specific_results, spend_file, [selected_year], "supplier") if scope31_supplier_specific_results is not None else None

    # Create columns in the new order: 1st Exiobase, 2nd Supplier-Specific, 3rd EPA
    scope31_col1, scope31_col2, scope31_col3 = st.columns(3)

    with scope31_col1:
        st.markdown("**Exiobase**")
        st.metric(
            "Country-Specific Factors", 
            f"{scope31_exiobase_total:,.2f}",
            help="A spend-based approach using country-specific emission factors from Exiobase."
        )
        st.caption("ton CO₂e", unsafe_allow_html=True)
        st.markdown("**Intensity**")
        intensity_text = f"{exiobase_intensity:,.1f} g CO₂e/$" if exiobase_intensity is not None else "N/A"
        st.markdown(intensity_text)

    with scope31_col2:
        st.markdown("**AKDN & Exiobase**")
        st.metric(
            "Supplier-Specific Factors", 
            f"{scope31_supplier_total:,.2f}",
            help="Combines AKDN supplier-specific emission factors with country-specific Exiobase fallback for most accurate estimation."
        )
        st.caption("ton CO₂e", unsafe_allow_html=True)
        st.markdown("**Intensity**")
        intensity_text = f"{supplier_intensity:,.1f} g CO₂e/$" if supplier_intensity is not None else "N/A"
        st.markdown(intensity_text)

    with scope31_col3:
        st.markdown("**EPA**")
        st.metric(
            "USA EPA Factors", 
            f"{scope31_total:,.2f}",
            help="A spend-based approach using USA EPA emission factors based on NAICS codes."
        )
        st.caption("ton CO₂e", unsafe_allow_html=True)
        st.markdown("**Intensity**")
        intensity_text = f"{epa_intensity:,.1f} g CO₂e/$" if epa_intensity is not None else "N/A"
        st.markdown(intensity_text)

    # Scope 3.4 section - TCE Breakdown
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("#### Scope 3.4 Emissions - Transportation Emissions")

    # Calculate Scope 3.4 intensity
    scope34_intensity = calculate_scope34_intensity(filtered_awb, filtered_bol)

    # Calculate subtotals for Air and Ocean emissions
    air_subtotal = (
        tce1_results.get('Air', 0.0) + 
        tce2_results.get('Air', 0.0) + 
        awb_emissions + 
        tce4_results.get('Air', 0.0)
    )

    ocean_subtotal = (
        tce1_results.get('Sea', 0.0) + 
        tce2_results.get('Sea', 0.0) + 
        bol_emissions + 
        tce4_results.get('Sea', 0.0)
    )

    scope34_grand_total = air_subtotal + ocean_subtotal

    # Calculate percentages for the speedometer
    air_percentage = (air_subtotal / scope34_grand_total * 100) if scope34_grand_total > 0 else 0
    ocean_percentage = (ocean_subtotal / scope34_grand_total * 100) if scope34_grand_total > 0 else 0

    # Create speedometer chart
    fig = go.Figure()

    # Add the speedometer gauge
    fig.add_trace(go.Indicator(
        mode="gauge",
        value=air_percentage,
        domain={'x': [0, 1], 'y': [0, 1]},
        title={'text': "Air vs Ocean Emissions", 'font': {'size': 24}},
        gauge={
            'axis': {'range': [0, 100]},
            'bar': {'color': "rgba(0,0,0,0)"},
            'steps': [
                {'range': [0, ocean_percentage], 'color': '#2D5889'},
                {'range': [ocean_percentage, 100], 'color': '#ED7136'}
            ],
            'bgcolor': "#f0f2f6"
        }
    ))

    fig.update_layout(
        height=400,
        margin=dict(l=50, r=50, t=100, b=50),
        font={'color': "darkblue", 'family': "Arial"},
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)'
    )

    # Grand Total above the chart
    st.markdown(f"""
    <div style='background-color: white; padding: 5px; border-radius: 10px; text-align: center; margin-bottom: 5px;'>
        <h2 style='margin: 0; color: #2c3e50;'>Scope 3.4 Total:  {scope34_grand_total:,.1f} ton CO₂e</h2>
        <h2 style='margin: 0; color: #2c3e50;'>Intensity:  {scope34_intensity:,.1f} g CO₂e/tkm</h2>
    </div>
    """, unsafe_allow_html=True)

    # Create columns for Ocean, Chart, and Air with bottom alignment
    ocean_col, chart_col, air_col = st.columns([1, 2, 1])
    
    with ocean_col:
        # Use container to help with bottom alignment
        ocean_container = st.container()
        with ocean_container:
            st.markdown(f"""
            <div style='text-align: center; padding: 15px; border-radius: 10px; background-color: white;'>
                <h1 style='margin: 0; color: white;'> </h1>
                <h2 style='margin: 0; color: #2D5889;'>Ocean</h2>
                <h2 style='margin: 0; color: #2D5889;'>{ocean_percentage:.1f}%</h2>
            </div>
            """, unsafe_allow_html=True)
        
        # Add spacer to push content to bottom
        st.markdown("<div style='height: 10px;'></div>", unsafe_allow_html=True)
    
    with chart_col:
        st.plotly_chart(fig, use_container_width=True)
    
    with air_col:
        # Use container to help with bottom alignment
        air_container = st.container()
        with air_container:
            st.markdown(f"""
            <div style='text-align: center; padding: 15px; border-radius: 10px; background-color: white;'>
                <h1 style='margin: 0; color: white;'> </h1>
                <h2 style='margin: 0; color: #ED7136;'>Air</h2>
                <h2 style='margin: 0; color: #ED7136;'>{air_percentage:.1f}%</h2>
            </div>
            """, unsafe_allow_html=True)
        
        # Add spacer to push content to bottom
        st.markdown("<div style='height: 10px;'></div>", unsafe_allow_html=True)

    # Add some spacing
    #st.markdown("<br>", unsafe_allow_html=True)

    # Create a grid for TCE 1-4 with Air, Ocean, and Total as row headers
    st.markdown("**Transport Chain Elements (TCE) Breakdown**")
    st.caption("Unit: ton CO₂e")

    # Create a DataFrame for the TCE data
    tce_data = {
        'Mode': ['Air', 'Ocean', 'Total'],
        'TCE 1': [
            f"{tce1_results.get('Air', 0.0):,.2f}",
            f"{tce1_results.get('Sea', 0.0):,.2f}",
            f"{tce1_results.get('Total', 0.0):,.2f}"
        ],
        'TCE 2': [
            f"{tce2_results.get('Air', 0.0):,.2f}",
            f"{tce2_results.get('Sea', 0.0):,.2f}",
            f"{tce2_results.get('Total', 0.0):,.2f}"
        ],
        'TCE 3': [
            f"{awb_emissions:,.2f}",
            f"{bol_emissions:,.2f}",
            f"{tce3_value:,.2f}"
        ],
        'TCE 4': [
            f"{tce4_results.get('Air', 0.0):,.2f}",
            f"{tce4_results.get('Sea', 0.0):,.2f}",
            f"{tce4_results.get('Total', 0.0):,.2f}"
        ],
        'Subtotal': [
            # Air subtotal (sum of TCE1 Air + TCE2 Air + TCE3 Air + TCE4 Air)
            f"{tce1_results.get('Air', 0.0) + tce2_results.get('Air', 0.0) + awb_emissions + tce4_results.get('Air', 0.0):,.2f}",
            # Ocean subtotal (sum of TCE1 Sea + TCE2 Sea + TCE3 Sea + TCE4 Sea)
            f"{tce1_results.get('Sea', 0.0) + tce2_results.get('Sea', 0.0) + bol_emissions + tce4_results.get('Sea', 0.0):,.2f}",
            # Total subtotal (sum of all TCEs)
            f"{tce1_results.get('Total', 0.0) + tce2_results.get('Total', 0.0) + tce3_value + tce4_results.get('Total', 0.0):,.2f}"
        ]
    }

    tce_df = pd.DataFrame(tce_data)

    # Style the DataFrame for better presentation
    styled_tce = tce_df.style \
        .set_table_styles([
            {'selector': 'thead th', 'props': [('background-color', '#f0f2f6'), 
                                            ('font-weight', 'bold'),
                                            ('text-align', 'center')]},
            {'selector': 'tbody tr:nth-child(even)', 'props': [('background-color', '#f9f9f9')]},
            {'selector': 'tbody tr:nth-child(odd)', 'props': [('background-color', 'white')]},
            {'selector': 'td', 'props': [('text-align', 'right'), ('padding', '8px')]},
            {'selector': 'th', 'props': [('text-align', 'left'), ('padding', '8px')]},
            # Highlight the subtotal column
            {'selector': 'td:nth-child(6), th:nth-child(6)', 'props': [('font-weight', 'bold')]}
        ]) \
        .hide(axis='index')

    # Display the table
    st.table(styled_tce)

    # Add tooltips/descriptions below the table
    st.caption("""
    **TCE Definitions:**
    - **TCE 1**: Inland Transport (to export hub)
    - **TCE 2**: Export Hub Operations  
    - **TCE 3**: Main Transport (air/ocean)
    - **TCE 4**: Import Hub Operations
    """)

    st.markdown("---")    
    st.markdown("<br>", unsafe_allow_html=True)
    st.header("Detailed Scope 3.1 and Scope 3.4 Analysis")

    # ==================== SCOPE 3.1 SUMMARY AND COMPARISON ====================
    if analysis_scope in ["Scope 3.1 (Purchased Goods)", "Both"]:
        if scope31_results is not None or scope31_exiobase_results is not None or scope31_supplier_specific_results is not None:
            st.subheader("Scope 3.1")
            
            # Add general explanation
            with st.expander("ℹ️ About Scope 3.1 Calculation Methods"):
                st.markdown("""
                **Scope 3.1: Purchased Goods and Services**
                
                Three different calculation methods are available:
                
                **1. EPA Method** 
                - A spend-based approach using standardized emission factors from the US Environmental Protection Agency
                - Emission factors are based on the NAICS codes of PAHO's product categories
                - Not country-specific, assumes US average supply chains
                
                **2. Exiobase Method** 
                - A spend-based approach using the Exiobase multi-regional input-output database
                - Accounts for country-specific emission factors based on suppliers' locations
                - Considers the countries and sectors involved in PAHO's supply chain
                
                **3. Supplier-Specific + Exiobase** 
                - A spend-based hybrid approach that prioritizes supplier-specific emission factors when available
                - Supplier-specific emission factors from Aga Khan Development Network are used first
                - Exiobase emission factors are used as fallback when supplier-specific data is unavailable
                """)
            
            # Combined visualization for all three methods
            if (scope31_results is not None and scope31_exiobase_results is not None and
                scope31_supplier_specific_results is not None and 
                len(scope31_results) > 1 and len(scope31_exiobase_results) > 1 and 
                len(scope31_supplier_specific_results) > 1):
                
                # Prepare comparison data
                epa_total = scope31_results[scope31_results['Supplier'] == 'TOTAL']['t CO2e'].values[0]
                exiobase_total = scope31_exiobase_results[
                    scope31_exiobase_results['Supplier'] == 'TOTAL'
                ]['Exiobase_t_CO2e'].values[0]
                supplier_total = scope31_supplier_specific_results[
                    scope31_supplier_specific_results['Supplier'] == 'TOTAL'
                ]['t_CO2e'].values[0]
                
                # Create comparison data in the desired order: 1st Exiobase, 2nd Supplier-Specific, 3rd EPA
                comparison_df = pd.DataFrame({
                    'Method': ['Exiobase', 'Supplier-Specific + Exiobase', 'EPA'],
                    'Emissions (tCO₂e)': [exiobase_total, supplier_total, epa_total]
                })
                
                fig = px.bar(
                    comparison_df,
                    x='Method',
                    y='Emissions (tCO₂e)',
                    color_discrete_sequence=['#C0C0C0'],  # Silver color for all bars
                    title='Total Emissions Comparison: All Methods',
                    labels={'Emissions (tCO₂e)': 'Total Emissions (tCO₂e)'},
                    text_auto='.1f'  # This automatically adds data labels with 1 decimal place
                )
                
                # Customize the layout
                fig.update_layout(
                    showlegend=False,  # Remove legend since we have clear labels
                    xaxis_title=None,  # Remove x-axis title
                    yaxis_title='Total Emissions (tCO₂e)',
                    font=dict(size=12),
                    plot_bgcolor='rgba(0,0,0,0)',
                    paper_bgcolor='rgba(0,0,0,0)'
                )
                
                # Improve data labels appearance
                fig.update_traces(
                    textposition='outside',  # Place data labels above the bars
                    textfont=dict(size=12, color='black'),
                    opacity=0.8
                )
                
                # Adjust y-axis to accommodate the data labels
                max_emission = max(exiobase_total, supplier_total, epa_total)
                fig.update_yaxes(range=[0, max_emission * 1.1])  # Add 10% padding for labels
                
                st.plotly_chart(fig, use_container_width=True)

    # ==================== SCOPE 3.4 SECTION ====================
    st.subheader("Scope 3.4")
    show_combined_map(filtered_awb, filtered_bol, precomputed_mappings)

    # Create columns for the new visualizations
    col1, col2, col3 = st.columns(3)

    with col1:
        # ==================== TOP 5 FLIGHT ROUTES BY COUNT ====================
        st.subheader("Top 5 Routes")
        
        top_routes = get_top_flight_routes(filtered_awb, n=5)
        
        if not top_routes.empty:
            fig_routes = px.bar(
                top_routes,
                x='flight_count',
                y='route',
                orientation='h',
                color_discrete_sequence=['#C0C0C0'],  # Single color instead of color scale
                labels={
                    'flight_count': 'Number of Flights',
                    'route': ''  # Remove Y-axis label
                },
                hover_data=['ghg_emissions_tCO2e', 'avg_distance_km'],
                title=''
            )
            fig_routes.update_layout(
                yaxis={'categoryorder': 'total ascending', 'title': ''},  # Remove Y-axis title
                showlegend=False,
                height=400
            )
            st.plotly_chart(fig_routes, use_container_width=True)
            
            # Display data table
            with st.expander("View Route Details"):
                display_df = top_routes.copy()
                display_df['Flights'] = display_df['flight_count']
                display_df['Emissions (tCO₂e)'] = display_df['ghg_emissions_tCO2e'].round(2)
                display_df['Avg Distance (km)'] = display_df['avg_distance_km'].round(0)
                st.dataframe(display_df[['route', 'Flights', 'Emissions (tCO₂e)', 'Avg Distance (km)']], 
                            hide_index=True)
        else:
            st.info("No flight route data available")

    with col2:
        # ==================== TOP 5 SOURCE LOCATIONS BY COUNT ====================
        st.subheader("Top 5 Origins")
        
        # Only pass AWB data, not BOL data
        top_sources = get_top_source_locations(filtered_awb, n=5)
        
        if not top_sources.empty:
            # Create stacked bar chart for shipment counts
            fig_sources = px.bar(
                top_sources,
                x='shipment_count',
                y='location',
                orientation='h',
                title='',
                color_discrete_sequence=['#C0C0C0'],
                labels={
                    'shipment_count': 'Number of Shipments',
                    'location': ''  # Remove Y-axis label
                },
                hover_data=['ghg_emissions_tCO2e']
            )
            fig_sources.update_layout(
                yaxis={'categoryorder': 'total ascending', 'title': ''},  # Remove Y-axis title
                height=400,
                showlegend=True
            )
            st.plotly_chart(fig_sources, use_container_width=True)
            
            # Display data table
            with st.expander("View Source Details"):
                display_df = top_sources.copy()
                display_df['Total Shipments'] = display_df['shipment_count']
                display_df['Total Emissions (tCO₂e)'] = display_df['ghg_emissions_tCO2e'].round(2)
                
                # Calculate shipment composition percentages
                if 'Air' in display_df.columns and 'Ocean' in display_df.columns:
                    display_df['Air %'] = (display_df['Air'] / display_df['shipment_count'] * 100).round(1)
                    display_df['Ocean %'] = (display_df['Ocean'] / display_df['shipment_count'] * 100).round(1)
                    st.dataframe(display_df[['location', 'Total Shipments', 'Air', 'Ocean', 'Air %', 'Ocean %', 'Total Emissions (tCO₂e)']], 
                                hide_index=True)
                else:
                    st.dataframe(display_df[['location', 'Total Shipments', 'Total Emissions (tCO₂e)']], 
                                hide_index=True)
        else:
            st.info("No AWB origin data available")

    with col3:
        # ==================== TOP 5 AIRLINES BY COUNT ====================
        st.subheader("Top 5 Airlines")
        
        if filtered_awb is not None and not filtered_awb.empty and 'airline' in filtered_awb.columns:
            # Prepare airline data - count flights instead of summing emissions
            airline_counts = filtered_awb[filtered_awb['leg'] != 'TOTAL'].groupby('airline').agg({
                'leg': 'count',
                'ghg_emissions_tCO2e': 'sum',
                'distance_km': 'sum'
            }).rename(columns={
                'leg': 'flight_count',
                'distance_km': 'total_distance_km'
            }).reset_index()
            
            # Get top 5 airlines by flight count
            top_airlines = airline_counts.nlargest(5, 'flight_count')
            
            if not top_airlines.empty:
                # Create bar chart
                fig = px.bar(
                    top_airlines,
                    x='flight_count',
                    y='airline',
                    orientation='h',
                    color_discrete_sequence=['#C0C0C0'],  # Single color instead of color scale
                    labels={
                        'flight_count': 'Number of Flights',
                        'airline': ''  # Remove Y-axis label
                    },
                    hover_data=['ghg_emissions_tCO2e', 'total_distance_km'],
                    title=''
                )
                fig.update_layout(
                    yaxis={'categoryorder': 'total ascending', 'title': ''},  # Remove Y-axis title
                    showlegend=False,
                    height=400
                )
                st.plotly_chart(fig, use_container_width=True)
                
                # Display data table
                with st.expander("View Airline Details"):
                    display_df = top_airlines.copy()
                    display_df['Flights'] = display_df['flight_count']
                    display_df['Emissions (tCO₂e)'] = display_df['ghg_emissions_tCO2e'].round(2)
                    display_df['Total Distance (km)'] = display_df['total_distance_km'].round(0)
                    st.dataframe(display_df[['airline', 'Flights', 'Emissions (tCO₂e)', 'Total Distance (km)']], 
                                hide_index=True)
            else:
                st.info("No airline data available")
        else:
            st.info("No airline data available")

    # ==================== ADDITIONAL METRICS ROW ====================
    st.subheader("Breakdown by Transport Chain Element (TCE)")
    st.caption("Unit: ton CO₂e")

    # Calculate TCE 1 (inland transport emissions)
    tce1_results = calculate_tce1_breakdown(filtered_awb, filtered_bol, precomputed_mappings)
    tce1_value = tce1_results['Total']

    # Calculate TCE 2 (export hub operations)
    tce2_results = calculate_tce2_breakdown(filtered_awb, filtered_bol)
    tce2_value = tce2_results['Total']

    # Calculate TCE 3 (same as sum of Air Transport Emissions and Ocean Transport Emissions)
    tce_3_value = awb_emissions + bol_emissions

    # Calculate TCE 4 (import hub operations)
    tce4_results = calculate_tce4_breakdown(filtered_awb, filtered_bol)
    tce4_value = tce4_results['Total']

    # Create metrics columns
    metric_col1, metric_col2, metric_col3, metric_col4 = st.columns(4)

    with metric_col1:
        st.metric("TCE 1", f"{tce1_value:,.1f}")

    with metric_col2:
        st.metric("TCE 2", f"{tce2_value:,.1f}")

    with metric_col3:
        st.metric("TCE 3", f"{tce_3_value:,.1f}")

    with metric_col4:
        st.metric("TCE 4", f"{tce4_value:,.1f}")

    # ==================== TABS DISPLAY ====================
    # Create tabs for all data tables
    tab_list = []
    
    if has_awb:
        tab_list.append("Air Waybill (AWB)")
    if has_bol:
        tab_list.append("Bill of Lading (BOL)")
    if analysis_scope in ["Scope 3.1 (Purchased Goods)", "Both"]:
        tab_list.extend(["EPA Method", "Exiobase Method", "Supplier-Specific + Exiobase"])
    
    if tab_list:
        tabs = st.tabs(tab_list)
        tab_index = 0
    else:
        tabs = []
        tab_index = 0

    # ==================== MODIFIED AWB TAB ====================
    if has_awb:
        with tabs[tab_index]:
            st.header("Air Waybill (AWB) Analysis")
            
            if awb_with_tce is not None and not awb_with_tce.empty:
                st.success(f"✅ Processed {len(awb_with_tce)} AWB records")
                
                st.subheader("Flight Data with TCE Breakdown")
                
                # Reorder columns to show TCE emissions at the end
                column_order = [
                    col for col in awb_with_tce.columns 
                    if col not in ['TCE1_Emissions_tCO2e', 'TCE2_Emissions_tCO2e', 'TCE4_Emissions_tCO2e']
                ] + ['TCE1_Emissions_tCO2e', 'TCE2_Emissions_tCO2e', 'TCE4_Emissions_tCO2e']
                
                styled_awb = style_awb_dataframe(awb_with_tce[column_order])
                st.dataframe(styled_awb)
                
                # Update download function to include TCE columns
                create_excel_download(awb_with_tce[column_order], "awb_emissions_with_tce.xlsx")

            else:
                st.info("No valid AWB data found" + (" (filtered out)" if selected_suppliers or selected_apos else ""))
        tab_index += 1

    # ==================== MODIFIED BOL TAB ====================
    if has_bol:
        with tabs[tab_index]:
            st.header("Bill of Lading (BOL) Analysis")
            
            if bol_with_tce is not None and not bol_with_tce.empty:
                st.success(f"✅ Processed {len(bol_with_tce)} BOL records")
                
                st.subheader("Shipping Data with TCE Breakdown")
                
                # Reorder columns to show TCE emissions at the end
                column_order = [
                    col for col in bol_with_tce.columns 
                    if col not in ['TCE1_Emissions_tCO2e', 'TCE2_Emissions_tCO2e', 'TCE4_Emissions_tCO2e']
                ] + ['TCE1_Emissions_tCO2e', 'TCE2_Emissions_tCO2e', 'TCE4_Emissions_tCO2e']
                
                st.dataframe(bol_with_tce[column_order])
                
                # Update download function to include TCE columns
                create_excel_download(bol_with_tce[column_order], "bol_emissions_with_tce.xlsx")
                
            else:
                st.error("❌ No valid BOL data found" + (" (filtered out)" if selected_suppliers or selected_apos else ""))
        tab_index += 1

    # ==================== SCOPE 3.1 DETAILED TABLES ====================
    if analysis_scope in ["Scope 3.1 (Purchased Goods)", "Both"]:
        # EPA Method Tab
        with tabs[tab_index]:
            st.header("EPA Method Details")
            
            if scope31_results is not None:
                formatted_df = scope31_results.copy()
                if 'Supplier' in formatted_df.columns:
                    formatted_df['Supplier'] = formatted_df['Supplier'].fillna('Unknown')
                
                st.dataframe(
                    formatted_df.style.format({
                        'Amount (USD)': '{:,.2f}',
                        'kg CO2e': '{:,.2f}',
                        't CO2e': '{:,.6f}',
                        'kg CO2e/USD': '{:,.6f}'
                    })
                )
                
                # Download button for EPA results
                create_excel_download(scope31_results, "scope31_epa_results.xlsx")
            else:
                st.info("No EPA results available")
        tab_index += 1
        
        # Exiobase Method Tab
        with tabs[tab_index]:
            st.header("Exiobase Method Details")
            
            if scope31_exiobase_results is not None:
                formatted_df = scope31_exiobase_results.copy()
                if 'Supplier' in formatted_df.columns:
                    formatted_df['Supplier'] = formatted_df['Supplier'].fillna('Unknown')
                
                st.dataframe(
                    formatted_df.style.format({
                        'Amount (USD)': '{:,.2f}',
                        'Exiobase_kg_CO2e': '{:,.2f}',
                        'Exiobase_t_CO2e': '{:,.6f}',
                        'Exiobase_kg_CO2e/USD': '{:,.6f}'
                    })
                )
                
                # Download button for Exiobase results
                create_excel_download(scope31_exiobase_results, "scope31_exiobase_results.xlsx")
            else:
                st.info("No Exiobase results available")
        tab_index += 1
        
        # Supplier-Specific + Exiobase Tab
        with tabs[tab_index]:
            st.header("Supplier-Specific + Exiobase Details")
            
            if scope31_supplier_specific_results is not None:
                formatted_df = scope31_supplier_specific_results.copy()
                if 'Supplier' in formatted_df.columns:
                    formatted_df['Supplier'] = formatted_df['Supplier'].fillna('Unknown')
                
                st.dataframe(formatted_df)
                
                # Download button for Supplier-Specific results
                create_excel_download(scope31_supplier_specific_results, "scope31_supplier_specific_results.xlsx")
            else:
                st.info("No Supplier-Specific results available")

if __name__ == "__main__":
    main()
